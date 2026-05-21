from shiny import App, Inputs, Outputs, Session, render, ui, reactive
import pandas as pd 
import io
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font,  Alignment, Color
import asyncio
import os
from shiny.types import ImgData
from datetime import datetime
import html
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors


img_folder = "www"
os.makedirs(img_folder, exist_ok=True)
local_image = os.path.join(img_folder, "Picture.png")  # Replace with your image file

# Glossary dictionary
glossary = {
    "AIV": ["Study code for IND enabling permeability studies", ""],
    "AUC": ["Area Under the plasma Concentration versus time curve extrapolated to infinity", ""],
    "BCRP": ["Breast Cancer Resistance Protein", ""],
    "Cgut": ["Concentration in the intestinal lumen", "Dose/250 mL"],
    "Cmax": ["Maximal systemic concentration in the plasma", ""],
    "Cmax,in": ["Estimated plasma maximal inhibitor concentration at the inlet to the liver", ""],
    "CYP": ["Cytochrome P450", ""],
    "DDI": ["Drug-Drug Interaction", ""],
    "d factor": ["An empirical calibration factor to enable in vitro-to-in vivo induction scaling", "Scaling factor from system used, otherwise assumed 1"],    
    "EC50": ["Concentration of inducer producing 50% of the maximum effect", ""],
    "Emax": ["Maximum fold induction", ""],
    "Fa": ["Fraction of dose absorbed at the gut level", "When Fa value is not available, a default value of 1 is assumed corresponding to entire absorption in gut wall for perpetrator drugs"],
    "Fg": ["Fraction of victim dose escaping metabolism within the walls of the gastrointestinal tract (intestinal wall availability)", "When Fg value is not available, a default value of 1 is assumed corresponding to no metabolism in gut wall for perpetrator drugs"],
    "fm": ["Fraction of hepatic metabolic clearance mediated by enzyme of interest", ""],
    "fu,p": ["Unbound fraction in plasma", "if fu,p < 0.01, keep fu,p=0.01"],
    "fu,mic": ["Unbound fraction in the microsomes, which is used to convert Ki, KI and IC50 values into corresponding unbound parameters", ""],
    "fu,hep": ["Unbound fraction in hepatocytes used to convert EC50 values from induction assays into corresponding unbound parameters", ""],
    "fu,inc": ["Unbound fraction in the assay system (if different from microsomes or hepatocytes)", ""],
    "HLM": ["Human Liver Microsomes", ""],
    "IC50": ["Half maximal inhibitory concentration", ""],
    "ICH": ["International Council for Harmonisation", ""],
    "ICH": ["Study code for IND enabling CYP inhibition studies", ""],
    "IHH": ["Study code for IND enabling CYP induction studies", ""],    
    "INT": ["DDI clinical interaction study", ""],
    "ka": ["Rate constant of absorption (/h) in humans", "If unknown use ka = 6 (worst case scenario)"],
    "Kdeg,g": ["Endogenous degradation rate constant of an enzyme in the intestine (gut)", ""],
    "Kdeg,h": ["Endogenous degradation rate constant of an enzyme in the liver (hepatic)", ""],
    "Ki": ["Dissociation constant of the enzyme-substrate complex, inhibition constant", ""],
    "KI": ["The concentration of SAR that produces one-half of kinact", ""],
    "kinact": ["Maximal rate of enzyme inactivation", ""],
    "Kobs": ["Observed inactivation rate constant of an affected enzyme", ""],
    "MATE": ["Multidrug And Toxin Extrusion transporter", ""],
    "MBI": ["Mechanism Based Inhibition", ""],
    "MW": ["Molecular Weight", ""],
    "OAT": ["Organic Anion Transporter", ""],
    "OATP": ["Organic Anion Transporting Polypeptide", ""],
    "OCT": ["Organic Cation Transporter", ""],
    "Pgp": ["P-glycoprotein efflux transporter", ""],
    "Qen": ["Intestinal blood flow", "Set by default at 18 L/h"],
    "Qh": ["Hepatic blood flow", "Set by default at 97 L/h"],
    "R": ["Interaction ratio (of the substrate drug with and without the perpetrator)", ""],
    "Rb": ["Blood-to-plasma ratio", "If unknown Rb = 1 for neutrals/bases or Rb = 0.55 for acids/zwitterions"],
    "TDI": ["Time Dependent Inhibition", ""],
    "TRI": ["Study code for IND enabling transporter inhibition studies", ""],
    "UGT": ["UDP-Glucuronosyl-Transferase", ""]
}

# Convert the glossary dictionary to a DataFrame
glossary_df = pd.DataFrame.from_dict(glossary, orient='index', columns=["Definition", "Note"])
glossary_df.reset_index(inplace=True)
glossary_df.rename(columns={'index': 'Term'}, inplace=True)

# List of all enzymes
all_enzymes_ei_unified = [
                        {"attribute": "Competitive enzyme inhibition", "name": "CYP1A2",  "id": "cyp1a2_ei",  "probe": "Phenacetin",      "intestinal": False},
                        {"attribute": "Competitive enzyme inhibition", "name": "CYP2B6",  "id": "cyp2b6_ei",  "probe": "Bupropion",        "intestinal": False},
                        {"attribute": "Competitive enzyme inhibition", "name": "CYP2C8",  "id": "cyp2c8_ei",  "probe": "Amodiaquine",      "intestinal": False},
                        {"attribute": "Competitive enzyme inhibition", "name": "CYP2C9",  "id": "cyp2c9_ei",  "probe": "Diclofenac",       "intestinal": False},
                        {"attribute": "Competitive enzyme inhibition", "name": "CYP2C19", "id": "cyp2c19_ei", "probe": "S-Mephenytoin",    "intestinal": False},
                        {"attribute": "Competitive enzyme inhibition", "name": "CYP2D6",  "id": "cyp2d6_ei",  "probe": "Dextromethorphan", "intestinal": False},
                        {"attribute": "Competitive enzyme inhibition", "name": "CYP3A4",   "id": "cyp3a4_ei",   "probe": "Midazolam",        "intestinal": True},
                        {"attribute": "Competitive enzyme inhibition", "name": "UGT1A1",  "id": "ugt1a1_ei",  "probe": "17β-Estradiol",    "intestinal": False},
                        {"attribute": "Competitive enzyme inhibition", "name": "UGT1A4",  "id": "ugt1a4_ei",  "probe": "Trifluoperazine",  "intestinal": False},
                        {"attribute": "Competitive enzyme inhibition", "name": "UGT1A9",  "id": "ugt1a9_ei",  "probe": "Propofol",         "intestinal": False},
                        {"attribute": "Competitive enzyme inhibition", "name": "UGT2B7",  "id": "ugt2b7_ei",  "probe": "Zidovudine",       "intestinal": False},
                        {"attribute": "Competitive enzyme inhibition", "name": "UGT2B15", "id": "ugt2b15_ei", "probe": "Oxazepam",         "intestinal": False},
]

all_enzymes_tdi = [
                        {"attribute":"Irreversible CYP inhibition - hepatic","name":"CYP1A2", "id":"cyp1a2_tdi", "probe":"Phenacetin", "kdeg_h":0.0180},
                        {"attribute":"Irreversible CYP inhibition - hepatic","name":"CYP2B6", "id":"cyp2b6_tdi", "probe":"Bupropion", "kdeg_h":0.0216},
                        {"attribute":"Irreversible CYP inhibition - hepatic","name":"CYP2C8", "id":"cyp2c8_tdi", "probe":"Amodiaquine", "kdeg_h":0.0318},
                        {"attribute":"Irreversible CYP inhibition - hepatic","name":"CYP2C9", "id":"cyp2c9_tdi", "probe":"Diclofenac", "kdeg_h":0.0066},
                        {"attribute":"Irreversible CYP inhibition - hepatic","name":"CYP2C19", "id":"cyp2c19_tdi", "probe":"S-Mephenytoin", "kdeg_h":0.0264},
                        {"attribute":"Irreversible CYP inhibition - hepatic","name":"CYP2D6", "id":"cyp2d6_tdi", "probe":"Dextromethorphan", "kdeg_h":0.0138},
                        {"attribute":"Irreversible CYP inhibition - hepatic","name":"CYP3A4", "id":"cyp3a4_tdi", "probe":"Midazolam", "kdeg_h":0.0192},
]

all_enzymes_ed = [
                        {"attribute":"Enzyme induction","name":"CYP1A2", "id":"cyp1a2_ed"},
                        {"attribute":"Enzyme induction","name":"CYP2B6", "id":"cyp2b6_ed"},
                        {"attribute":"Enzyme induction","name":"CYP3A4", "id":"cyp3a4_ed"},
                        {"attribute":"Enzyme induction","name":"CYP2C8", "id":"cyp2c8_ed"},
                        {"attribute":"Enzyme induction","name":"CYP2C9", "id":"cyp2c9_ed"}, 
                        {"attribute":"Enzyme induction","name":"CYP2C19", "id":"cyp2c19_ed"}, 
                        {"attribute":"Enzyme induction","name":"UGT1A1", "id":"ugt1a1_ed"}, 
                        {"attribute":"Enzyme induction","name":"UGT1A4", "id":"ugt1a4_ed"}
]

all_enzymes_ne_unified = [
                        {"attribute": "Net effect", "name": "CYP1A2",  "id": "cyp1a2_ne",  "substrate": "Caffeine",        "probe": "Phenacetin",       "fm": 0.80, "fg": 1.0},
                        {"attribute": "Net effect", "name": "CYP2B6",  "id": "cyp2b6_ne",  "substrate": "Bupropion",       "probe": "Bupropion",        "fm": 0.85, "fg": 1.0},
                        {"attribute": "Net effect", "name": "CYP3A4",  "id": "cyp3a4_ne",  "substrate": "Midazolam",       "probe": "Midazolam",        "fm": 0.93, "fg": 0.44},
                        {"attribute": "Net effect", "name": "CYP2C8",  "id": "cyp2c8_ne",  "substrate": "Repaglinide",     "probe": "Amodiaquine",      "fm": 0.80, "fg": 1.0},
                        {"attribute": "Net effect", "name": "CYP2C9",  "id": "cyp2c9_ne",  "substrate": "Warfarin",        "probe": "Diclofenac",       "fm": 0.90, "fg": 1.0},
                        {"attribute": "Net effect", "name": "CYP2C19", "id": "cyp2c19_ne", "substrate": "Omeprazole",      "probe": "S-Mephenytoin",    "fm": 0.85, "fg": 1.0},
                        {"attribute": "Net effect", "name": "CYP2D6",  "id": "cyp2d6_ne",  "substrate": "Desipramine",     "probe": "Dextromethorphan", "fm": 0.80, "fg": 1.0},
                        {"attribute": "Net effect", "name": "UGT1A1",  "id": "ugt1a1_ne",  "substrate": "Cabotegravir",     "probe": "17β-Estradiol",    "fm": 0.54, "fg": 0.85},
                        {"attribute": "Net effect", "name": "UGT1A4",  "id": "ugt1a4_ne",  "substrate": "Trifluoperazine", "probe": "Trifluoperazine",  "fm": 0.75, "fg": 1.0},
]

all_transporters_ie = [
                        {"attribute":"Transporter inhibition - intestinal efflux","name":"Pgp", "id":"pgp_ti_ie"},
                        {"attribute":"Transporter inhibition - intestinal efflux","name":"BCRP", "id":"bcrp_ti_ie"},
]

all_transporters_hu = [
                        {"attribute":"Transporter inhibition - hepatic uptake","name":"OATP1B1", "id":"oatp1b1_ti_hu"},
                        {"attribute":"Transporter inhibition - hepatic uptake","name":"OATP1B3", "id":"oatp1b3_ti_hu"},
                        {"attribute":"Transporter inhibition - hepatic uptake","name":"OCT1", "id":"oct1_ti_hu"},
]

all_transporters_ru = [
                        {"attribute":"Transporter inhibition - renal uptake","name":"OAT1", "id":"oatp1b1_ti_ru"},
                        {"attribute":"Transporter inhibition - renal uptake","name":"OAT3", "id":"oatp1b3_ti_ru"},
                        {"attribute":"Transporter inhibition - renal uptake","name":"OCT2", "id":"oct1_ti_ru"},
]

all_transporters_re = [
                        {"attribute":"Transporter inhibition - renal efflux","name":"MATE1", "id":"mate1_ti_re"},
                        {"attribute":"Transporter inhibition - renal efflux","name":"MATE2K", "id":"mate2k_ti_re"},
                        {"attribute":"Transporter inhibition - renal efflux","name":"Pgp", "id":"pgp_ti_re"},
                        {"attribute":"Transporter inhibition - renal efflux","name":"BCRP", "id":"bcrp_ti_re"},                           
]

# Define the UI
app_ui = ui.page_fluid(

     ui.tags.style(
         

            """
        /* Style for Navset Pill List */
        .nav-pills .nav-link.active {
            background-color: #4B0082 !important; /* Dark Purple for Active Tab */
            color: white !important;
            font-size: 25px !important; /* Increase font size */
            padding: 12px 20px !important; /* Increase padding */
        }
        
        .nav-pills .nav-link {
            background-color: #D8BFD8 !important; /* Light Purple for Inactive Tabs */
            color: black !important;
        }


        /* Centering Title */
        .app-title {
            text-align: center;
            font-size: 24px;
            font-weight: bold;
            margin-bottom: 20px;
        }
    """,

            """
            .custom-title {
                font-size: 30px;
                color: white;
                background-color: #23004C;
                padding: 20px;
                border-radius: 5px;
                text-align: center;
                width: 100%;
                box-sizing: border-box;
                margin-bottom: 0px;
            }

         .custom-calculate-button {
            background-color: #4CAF50;  /* Green background */
            color: white;               /* White text */
            border: none;               /* Remove border */
            padding: 10px 40px;         /* Add some padding */
            text-align: center;         /* Center text */
            text-decoration: none;      /* Remove underline from the text */
            display: inline-block;      /* Keep it inline */
            font-size: 16px;            /* Increase font size */
            border-radius: 10px;         /* Add rounded corners */
            cursor: pointer;            /* Add pointer on hover */
            width: 200px;
        }
        .custom-calculate-button:hover {
            background-color: #45a049;  /* Darker green on hover */
        }

        .custom-summary-button {
            background-color: #4CAF50;  /* Green background */
            color: white;               /* White text */
            border: none;               /* Remove border */
            padding: 10px 40px;         /* Add some padding */
            text-align: center;         /* Center text */
            text-decoration: none;      /* Remove underline from the text */
            display: inline-block;      /* Keep it inline */
            font-size: 16px;            /* Increase font size */
            border-radius: 10px;         /* Add rounded corners */
            cursor: pointer;            /* Add pointer on hover */
            width: 300px;
        }

        .custom-download-button {
            background-color:rgb(78, 76, 175);  /* Blue background */
            color: white;               /* White text */
            border: none;               /* Remove border */
            padding: 10px 40px;         /* Add some padding */
            text-align: center;         /* Center text */
            text-decoration: none;      /* Remove underline from the text */
            display: inline-block;      /* Keep it inline */
            font-size: 16px;            /* Increase font size */
            border-radius: 10px;         /* Add rounded corners */
            cursor: pointer;            /* Add pointer on hover */
            width: 550px;
        }

        .custum-table {
                        text-align: left;
                        width: 100%;
                        border-collapse: collapse;
        }
        
        .custum-table th {
                        padding: 10px;
                        border: 1px solid #ddd;
                        background-color: #f2f2f2;
                        text-align: left;
        }

        .custum-table td {
                        padding: 10px;
                        border: 1px solid #ddd;
        }   
            """
        ),

        #App Title
        ui.tags.div(
            ui.tags.strong("Static DDI Risk Calculator"), class_="custom-title"
                ),

        ui.row( 

            ui.column(2,
            ui.div(
            ui.div(
            ui.h2(
                    " ", 
             ),

            ui.div(
            
                ),
            ui.input_text("cmp", "Compound", value="SARXXXXXX"),
            ui.input_numeric("mw", "MW (g/mol)", value=300, step=10),
            ui.input_numeric("cmax", "Cmax", value=500, step=10),
            ui.input_radio_buttons("cmax_unit", "", {
                "ng_ml": "ng/mL",
                "umol_l": "μmol/L"
            }, inline=True, selected="ng_ml"),
            ui.input_numeric("fup", "fu,p", min=0, max=1, value=0.01, step=0.001),
                )
            )
            ),

            ui.column(2,
            ui.div(
            ui.div(
            ui.h2(
                    " ", 
             ),

            ui.div(
            
                ),
            ui.input_numeric("dose", "Dose (mg)", value=200, step=1),
            ui.input_numeric("ka", "ka (/h)", value=6, step=0.1),
            ui.input_numeric("fa", "Fa", value=1, step=0.1, max=1, min=0),
            ui.input_numeric("fg", "Fg", value=1, step=0.1, max=1, min=0), 
                )
            )
            ),

            ui.column(2,
            ui.div(
            ui.div(
            ui.h2(
                    " ", 
             ),

            ui.div(
            
                ),
            # fa, fg, ka, Qh, Rb Input

            ui.input_numeric("qh", "Qh (L/h)", value=97, step=0.1),
            ui.input_numeric("rb", "Rb", value=0.55, step=0.01),
            ui.input_text_area("user_memo", "User memo", value="", height="120px")
                )
            )
            ),

             # Right side: static picture
    ui.column(
        6,
         ui.div(  # Wrapper div to center the image
        ui.output_image("image"),
        style="text-align: center;"  # Centers the image horizontally
    ),
        
    )
            ),
      
    ui.navset_pill_list(  
    ui.nav_panel(
       "Enzyme Inhibition",
        ui.tags.div(
            ui.tags.p("Basic Static Model", 
                     style="font-size: 14px; color: white; background-color: #4B0082; padding: 8px; border-radius: 4px;"),
        ),
      
    #Enzyme inhibition
    ui.div(
    ui.card(
        ui.card_header("Enzyme inhibition", style="font-size: 24px;"),  # Card Header
            
        # fumic Input
          ui.input_select("fumic", "fu,mic:", {
            "input_fumic": "User input",
            "austin_fumic": "Method of Austin",
            "hallifax_fumic": "Method of Hallifax"
        }),
        ui.output_ui("input_ui_fumic"),
        ui.output_text_verbatim("result_fumic"),

    #Competitive inhibition - hepatic + intestinal 
        ui.div(
        ui.card(
            ui.h2(
                "Competitive inhibition",
                ui.span("+", class_="collapseSymbol", style="float: right;"),
                href="#collapseCIH",
                class_="collapse-toggle collapse-title",
                data_bs_toggle="collapse",
                role="button",
                aria_expanded="false",
                aria_controls="collapseCIH",
                style="font-size: 24px; font-weight: bold;"
            ),
            ui.div(
                ui.div(
                    ui.card(
                        ui.card_header("Parameters"),
                        ui.card_body(
                            ui.h5("Select Enzymes"),
                            ui.output_ui("enzyme_selection_ui_ei"), 
                            ui.output_ui("custom_ei_inputs_container"),
                            ui.div(
                                ui.input_action_button(
                                    "add_custom_ei_button", 
                                    "+ Add custom enzyme",
                                    class_="btn btn-outline-primary btn-sm"
                                ),
                                style="margin-top: 10px;"
                            ),
                        )
                    )
                ),
                # Single calculate button for both hepatic and intestinal
                ui.div(
                    ui.input_action_button("calculate_all_ei", "Calculate", class_="custom-calculate-button")
                ),
                # Single result table showing both hepatic and intestinal results
                ui.div(
                    ui.card(
                        ui.card_header("Results (Competitive inhibition - hepatic & intestinal)"),
                        ui.card_body(
                            ui.output_table("result_table_ei_unified", class_="custum-table"),
                            ui.tags.head(
                                ui.tags.script(src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js")
                            ),
                            ui.tags.div(
                            "$$R_{\\text{hepatic}} = \\frac{C_{\\text{max}}[\\mu M] \\cdot f_{\\text{u,p}}}{K_{\\text{i}}[\\mu M] \\cdot f_{\\text{u,mic}}}$$",
                            "$$R_{\\text{intestinal (CYP3A4 only)}} = \\frac{Dose[mg] \\cdot 1000000}{MW[g/mol] \\cdot 250[mL] \\cdot K_{\\text{i}}[\\mu M] \\cdot f_{\\text{u,mic}}}$$",
                            style="text-align: center; font-size: 1.2em; margin-top: 20px;"
                            ),
                            ui.tags.footer("Hepatic: Red and bold if R >0.02, Green if R =<0.02"),
                            ui.tags.footer("Intestinal (CYP3A4): Red and bold if R >10, Green if R =<10"),
                        )
                    )
                ),
                class_="collapse", id="collapseCIH",
            ),
        ),
        ),  # Competitive inhibition unified card end

    ui.div(
    ui.card(
        ui.h2(
             "Irreversible CYP inhibition - hepatic",
                    ui.span("+", class_="collapseSymbol", style="float: right;"),
                    href="#collapseIIH",
                    class_="collapse-toggle collapse-title",
                    data_bs_toggle="collapse", 
                    role="button", 
                    aria_expanded="false", 
                    aria_controls="collapseIIH",
                    style="font-size: 24px; font-weight: bold;"
                ),
    ui.div(
    #ui.card_header("Irreversible CYP inhibition - hepatic", style="font-size: 24px;"),  # Card Header
   # Parameters Card
    ui.div(
        ui.card(
            ui.card_header("Parameters"),  # Card Header         
            ui.card_body(
                # Enzyme Selection and Parameters Subsection
                ui.h5("Select Enzymes"),
                ui.output_ui("enzyme_selection_ui_tdi"), 
                ui.output_ui("custom_tdi_inputs_container"),
                ui.div(
                    ui.input_action_button(
                        "add_custom_tdi_button", 
                        "+ Add custom enzyme",
                        class_="btn btn-outline-primary btn-sm"
                    ),
                    style="margin-top: 10px;"
                ),
            ),#parameters end
        )
    ),

    # Calculate Card
    ui.div(
            ui.input_action_button("calculate_all_tdi", "Calculate",class_="custom-calculate-button")  # Calculate Button    
        ),

    # Results Card
    ui.div(
        ui.card(
            ui.card_header("Results (Irreversible CYP inhibition - hepatic)"),  # Card Header
            ui.card_body(
               ui.output_table("result_table_tdi", class_="custum-table"),  # Display results in a table

                    # Display LaTeX Equation
                ui.tags.head(
                    ui.tags.script(src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js")
                    ),
                ui.tags.div(
                    "$$R_{\\text{hepatic}} = \\frac{K_{\\text{obs}}[/h] + K_{\\text{deg,h}}[/h]}{K_{\\text{deg,h}}[/h]}$$",
                    "$$\\text{where } K_{\\text{obs}}[/h] = \\frac{K_{\\text{inact}}[/h] \\cdot 5 \\cdot C_{\\text{max}}[\\mu M] \\cdot f_{\\text{u,p}}}{K_{\\text{I}}[\\mu M] \\cdot f_{\\text{u,mic}} + 5 \\cdot C_{\\text{max}}[\\mu M] \\cdot f_{\\text{u,p}}}$$",
                    style="text-align: center; font-size: 1.2em; margin-top: 20px;"
                ),
                ui.tags.footer("Red and bold if R >1.25, Green if R =<1.25")
            )
        )
    ),
    class_="collapse", id="collapseIIH",
    ),
    ),#Irreversible CYP inhibition - hepatic card end
    ),#Irreversible CYP inhibition - hepatic div end

    )#Enzyme inhibition card end
    )#Enzyme inhibition div end
    ),#Enzyme inhibition panel end


    ui.nav_panel(
        "Enzyme Induction",
        ui.tags.div(
            ui.tags.p("Basic Static Model", 
                     style="font-size: 14px; color: white; background-color: #4B0082; padding: 8px; border-radius: 4px;"),
        ),
    # Parameters Card
    ui.div(
        ui.card(
            ui.card_header("Enzyme Induction", style="font-size: 24px;"),  # Card Header
                ui.card_body(
                    # d_factor Input
                    ui.input_numeric("d_factor", "d factor", value=1, step=0.1),
                    
                    # fuhep Input
                    ui.input_select("fuhep", "fu,hep:", {
                    "input_fuhep": "User input",
                    "austin_fuhep": "Method of Austin",
                    "kilfold_fuhep": "Method of Kilfold"
                    }),
                    ui.output_ui("input_ui_fuhep"),
                    ui.output_text_verbatim("result_fuhep"),
                    
                    # Enzyme Selection and Parameters Subsection
                    ui.h5("Select Enzymes:"),
                    ui.output_ui("enzyme_selection_ui_ed"), 
                    ui.output_ui("custom_ed_inputs_container"),
                    ui.div(
                        ui.input_action_button(
                            "add_custom_ed_button", 
                            "+ Add custom enzyme",
                            class_="btn btn-outline-primary btn-sm"
                        ),
                        style="margin-top: 10px;"
                    ),
                )
        )
    ),

    # Calculate Card
    ui.div(   
        ui.input_action_button("calculate_all_ed", "Calculate",class_="custom-calculate-button")  # Calculate Button         
    ),

    # Results Card
    ui.div(
        ui.card(
            ui.card_header("Results (Enzyme induction)"),  # Card Header
            ui.card_body(
                ui.output_table("result_table_ed", class_="custum-table"),  # Display results in a table

                    # Display LaTeX Equation
                ui.tags.head(
                    ui.tags.script(src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js")
                    ),
                ui.tags.div(
                    "$$R_{\\text{induction}} = \\frac{1}{1 + d \\cdot \\frac{E_{\\text{max}} \\cdot 10 \\cdot C_{\\text{max}}[\\mu M] \\cdot f_{\\text{u,p}}}{EC_{50}[\\mu M] \\cdot f_{\\text{u,hep}} + 10 \\cdot C_{\\text{max}}[\\mu M] \\cdot f_{\\text{u,p}}}}$$",
                    style="text-align: center; font-size: 1.2em; margin-top: 20px;"
                ),
                ui.tags.footer("Red if R <0.2, Orange if 0.2=< R <0.5, Blue if 0.5=< R <0.8, Green if R >= 0.8") 
            )
        )
    )
),# Enzyme induction panel end


ui.nav_panel(
        "Transporter Inhibition",
        ui.tags.div(
            ui.tags.p("Basic Static Model", 
                     style="font-size: 14px; color: white; background-color: #4B0082; padding: 8px; border-radius: 4px;"),
        ),

    #Transporter inhibition
    ui.div(
    ui.card(
        ui.card_header("Transporter inhibition", style="font-size: 24px;"),  # Card Header
        ui.input_numeric("fuinc", "fu,inc", value=0.1, step=0.01),
    ui.div(
    ui.card(
        ui.h2(
            "Transporter inhibition - intestinal efflux",
                    ui.span("+", class_="collapseSymbol", style="float: right;"),
                    href="#collapseTIIE",
                    class_="collapse-toggle collapse-title",
                    data_bs_toggle="collapse", 
                    role="button", 
                    aria_expanded="false", 
                    aria_controls="collapseTIIE",
                    style="font-size: 24px; font-weight: bold;"
                ),

            
            ui.div(
       
   # Parameters Card
    ui.div(
        ui.card(
            ui.card_header("Parameters"),  # Card Header         
            ui.card_body(
                # Transporter Selection and Parameters Subsection
                ui.h5("Select Transporters"),               
                ui.output_ui("transporter_selection_ui_ti_ie"),         
                ui.output_ui("custom_ti_ie_inputs_container"),
                ui.div(
                    ui.input_action_button(
                        "add_custom_ti_ie_button", 
                        "+ Add custom transporter",
                        class_="btn btn-outline-primary btn-sm"
                    ),
                    style="margin-top: 10px;"
                ),
            )
        )
    ),

    # Calculate Card
    ui.div(
        ui.input_action_button("calculate_all_ti_ie", "Calculate",class_="custom-calculate-button")  # Calculate Button
    ),

    # Results Card
    ui.div(
        ui.card(
            ui.card_header("Results (Transporter inhibition - intestinal efflux)"),  # Card Header
            ui.card_body(
               ui.output_table("result_table_ti_ie", class_="custum-table"),  # Display results in a table

                    # Display LaTeX Equation
                ui.tags.head(
                    ui.tags.script(src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js")
                    ),
                ui.tags.div(
                    "$$R_{\\text{intestinal}} = \\frac{Dose[mg] \\cdot 1000000}{MW[g/mol] \\cdot 250[mL] \\cdot IC_{50}[\\mu M] \\cdot f_{\\text{u,inc}}}$$",
                    style="text-align: center; font-size: 1.2em; margin-top: 20px;"
                ),
                ui.tags.footer("Red and bold if R >10, Green if R =<10")
            )
        )
    ),
    class_="collapse", id="collapseTIIE",
    ),
    ),#Transporter inhibition - intestinal efflux card end
    ),#Transporter inhibition - intestinal efflux div end


    ui.div(
    ui.card(
        ui.h2(
            "Transporter inhibition - hepatic uptake",
                    ui.span("+", class_="collapseSymbol", style="float: right;"),
                    href="#collapseTIHU",
                    class_="collapse-toggle collapse-title",
                    data_bs_toggle="collapse", 
                    role="button", 
                    aria_expanded="false", 
                    aria_controls="collapseTIHU",
                    style="font-size: 24px; font-weight: bold;"
                ),

            
            ui.div(
        
   # Parameters Card
    ui.div(
        ui.card(
            ui.card_header("Parameters"),  # Card Header         
            ui.card_body(
                # Transporter Selection and Parameters Subsection
                ui.h5("Select Transporters"),
                ui.output_ui("transporter_selection_ui_ti_hu"),         
                ui.output_ui("custom_ti_hu_inputs_container"),
                ui.div(
                    ui.input_action_button(
                        "add_custom_ti_hu_button", 
                        "+ Add custom transporter",
                        class_="btn btn-outline-primary btn-sm"
                    ),
                    style="margin-top: 10px;"
                ),             
            )
        )
    ),

    # Calculate Card
    ui.div(
       ui.input_action_button("calculate_all_ti_hu", "Calculate",class_="custom-calculate-button")  # Calculate Button
    ),

    # Results Card
    ui.div(
        ui.card(
            ui.card_header("Results (Transporter inhibition - hepatic uptake)"),  # Card Header
            ui.card_body(
                ui.output_table("result_table_ti_hu", class_="custum-table"),  # Display results in a table

                    # Display LaTeX Equation
                ui.tags.head(
                    ui.tags.script(src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js")
                    ),    
                ui.tags.div(
                    "$$R_{\\text{hepatic}} = \\frac{C_{\\text{max,inlet}}[\\mu M] \\cdot f_{\\text{u,p}}}{IC_{50}[\\mu M] \\cdot f_{\\text{u,inc}}}$$",
                    "$$\\text{where } C_{\\text{max,inlet}}[\\mu M] = C_{\\text{max}}[\\mu M] + \\frac{F_a \\cdot F_g \\cdot k_a[/h] \\cdot \\text{Dose}[mg] \\cdot 1000}{MW[g/mol] \\cdot Q_h[L/h] \\cdot R_B}$$",
                    style="text-align: center; font-size: 1.2em; margin-top: 20px;"
                ),
                ui.tags.footer("Red and bold if R >0.1, Green if R =<0.1")  
            )
        )
    ),
        class_="collapse", id="collapseTIHU",
    ),
    ),#Transporter inhibition - hepatic uptake card end
    ),#Transporter inhibition - hepatic uptake div end


    ui.div(
    ui.card(
        ui.h2(
            "Transporter inhibition - renal uptake",
                    ui.span("+", class_="collapseSymbol", style="float: right;"),
                    href="#collapseTIRU",
                    data_bs_toggle="collapse", 
                    role="button", 
                    aria_expanded="false", 
                    aria_controls="collapseTIRU",
                    style="font-size: 24px; font-weight: bold;"
                ),

            
            ui.div(
       
   # Parameters Card
    ui.div(
        ui.card(
            ui.card_header("Parameters"),  # Card Header         
            ui.card_body(
                # Transporter Selection and Parameters Subsection
                ui.h5("Select Transporters"),   
                ui.output_ui("transporter_selection_ui_ti_ru"),         
                ui.output_ui("custom_ti_ru_inputs_container"),
                ui.div(
                    ui.input_action_button(
                        "add_custom_ti_ru_button", 
                        "+ Add custom transporter",
                        class_="btn btn-outline-primary btn-sm"
                    ),
                    style="margin-top: 10px;"
                ),                       
            )
        )
    ),
    # Calculate Card
    ui.div(
         ui.input_action_button("calculate_all_ti_ru", "Calculate",class_="custom-calculate-button")  # Calculate Button   
    ),

    # Results Card
    ui.div(
        ui.card(
            ui.card_header("Results (Transporter inhibition - renal uptake)"),  # Card Header
            ui.card_body(
                ui.output_table("result_table_ti_ru", class_="custum-table"),  # Display results in a table

                    # Display LaTeX Equation
                ui.tags.head(
                    ui.tags.script(src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js")
                    ),
                ui.tags.div(
                    "$$R_{\\text{renal}} = \\frac{C_{\\text{max}}[\\mu M] \\cdot f_{\\text{u,p}}}{IC_{50}[\\mu M] \\cdot f_{\\text{u,inc}}}$$",
                    style="text-align: center; font-size: 1.2em; margin-top: 20px;"
                ),
                ui.tags.footer("Red and bold if R >0.1, Green if R =<0.1")
            )
        )
    ),
     class_="collapse", id="collapseTIRU",
    ),
    ),#Transporter inhibition - renal uptake card end
    ),#Transporter inhibition - renal uptake div end


    ui.div(
    ui.card(
        ui.h2(
            "Transporter inhibition - renal efflux",
                    ui.span("+", class_="collapseSymbol", style="float: right;"),
                    href="#collapseTIRE",
                    data_bs_toggle="collapse", 
                    role="button", 
                    aria_expanded="false", 
                    aria_controls="collapseTIRE",
                    style="font-size: 24px; font-weight: bold;"
                ),

            
            ui.div(
        # Card Header
   # Parameters Card
    ui.div(
        ui.card(
            ui.card_header("Parameters"),  # Card Header         
            ui.card_body(
                # Transporter Selection and Parameters Subsection
                ui.h5("Select Transporters"),   
                ui.output_ui("transporter_selection_ui_ti_re"),         
                ui.output_ui("custom_ti_re_inputs_container"),
                ui.div(
                    ui.input_action_button(
                        "add_custom_ti_re_button", 
                        "+ Add custom transporter",
                        class_="btn btn-outline-primary btn-sm"
                    ),
                    style="margin-top: 10px;"
                ),                       
            )
        )
    ),
    # Calculate Card
    ui.div(
        ui.input_action_button("calculate_all_ti_re", "Calculate",class_="custom-calculate-button")  # Calculate Button
    ),

    # Results Card
    ui.div(
        ui.card(
            ui.card_header("Results (Transporter inhibition - renal efflux)"),  # Card Header
            ui.card_body(
                ui.output_table("result_table_ti_re", class_="custum-table"),  # Display results in a table

                    # Display LaTeX Equation
                ui.tags.head(
                    ui.tags.script(src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js")
                    ),
                ui.tags.div(
                    "$$R_{\\text{renal}} = \\frac{C_{\\text{max}}[\\mu M] \\cdot f_{\\text{u,p}}}{IC_{50}[\\mu M] \\cdot f_{\\text{u,inc}}}$$",
                    style="text-align: center; font-size: 1.2em; margin-top: 20px;"
                ),
                ui.tags.footer("Red and bold if R >0.02, Green if R =<0.02")
            )
        )
    ),
     class_="collapse", id="collapseTIRE",
    ),
    ),#Transporter inhibition - renal efflux card end
    ),#Transporter inhibition - renal efflux div end

    )#Transporter inhibition card end
    )#Transporter inhibition div end
), # Transporter inhibition panel end



    ui.nav_panel(
        "Net Effect (AUCR)",
        ui.tags.div(
            ui.tags.p("Mechanistic Static Model", 
                     style="font-size: 14px; color: white; background-color: #006400; padding: 8px; border-radius: 4px;"),
        ),
    # Parameters Card
    ui.div(
        ui.card(
            ui.card_header("Net effect", style="font-size: 24px;"),  # Card Header
            ui.card_body(              
                ui.input_numeric("qen", "Qen (L/h)", value=18, step=0.1),
                ui.h5("Select Enzymes:"),
                ui.output_ui("enzyme_selection_ui_ne"),
                ui.output_ui("custom_ne_inputs_container"),
                ui.div(
                    ui.input_action_button(
                        "add_custom_ne_button", 
                        "+ Add custom enzyme",
                        class_="btn btn-outline-primary btn-sm"
                    ),
                    style="margin-top: 10px;"
                ), 
            )
        )
    ),

    # Calculate Card
    ui.div(   
        ui.input_action_button("calculate_all_ne", "Calculate",class_="custom-calculate-button")  # Calculate Button         
    ),

    # Results Card
    ui.div(
        ui.card(
            ui.card_header("Results (Net effect)"),  # Card Header
            ui.card_body(
                ui.output_table("result_table_ne_unified", class_="custum-table"),  # Display results in a table

                    # Display LaTeX Equation
                ui.tags.head(
                    ui.tags.script(src="https://cdn.jsdelivr.net/npm/mathjax@3/es5/tex-mml-chtml.js")
                    ),
                ui.tags.div(
                    "$$AUCR (R) = \\frac{1}{[A_{\\text{g}} \\cdot B_{\\text{g}} \\cdot C_{\\text{g}}] \\cdot (1 - F_{\\text{g,victim}}) + F_{\\text{g,victim}}} \\cdot \\frac{1}{[A_{\\text{h}} \\cdot B_{\\text{h}} \\cdot C_{\\text{h}}] \\cdot f_{\\text{m}} + (1 - f_{\\text{m}})}$$",
                    "$$A_{\\text{g}} = \\frac{1}{1 + \\frac{[I]_{\\text{g}}[\\mu M]}{K_{\\text{i}}[\\mu M] \\cdot f_{\\text{u,mic}}}}, B_{\\text{g}} = \\frac{K_{\\text{deg,g}}}{K_{\\text{deg,g}} + \\frac{[I]_{\\text{g}}[\\mu M] \\cdot K_{\\text{inact}}}{[I]_{\\text{g}}[\\mu M] + K_{\\text{I}}[\\mu M] \\cdot f_{\\text{u,mic}}}}, C_{\\text{g}} = 1 + \\frac{d \\cdot E_{\\text{max}} \\cdot [I]_{\\text{g}}[\\mu M]}{[I]_{\\text{g}}[\\mu M] + EC_{\\text{50}}[\\mu M] \\cdot f_{\\text{u,hep}}}$$",
                    "$$A_{\\text{h}} = \\frac{1}{1 + \\frac{[I]_{\\text{h}}[\\mu M]}{K_{\\text{i}}[\\mu M] \\cdot f_{\\text{u,mic}}}}, B_{\\text{h}} = \\frac{K_{\\text{deg,h}}}{K_{\\text{deg,h}} + \\frac{[I]_{\\text{h}}[\\mu M] \\cdot K_{\\text{inact}}}{[I]_{\\text{h}}[\\mu M] + K_{\\text{I}}[\\mu M] \\cdot f_{\\text{u,mic}}}}, C_{\\text{h}} = 1 + \\frac{d \\cdot E_{\\text{max}} \\cdot [I]_{\\text{h}}[\\mu M]}{[I]_{\\text{h}}[\\mu M] + EC_{\\text{50}}[\\mu M] \\cdot f_{\\text{u,hep}}}$$",
                    "$$[I]_{\\text{g}}[\\mu M] = F_{\\text{a}} \\cdot k_{\\text{a}} \\cdot \\frac{Dose[mg] \\cdot 1000}{MW[g/mol] \\cdot Q_{\\text{en}}}, \\quad [I]_{\\text{h}}[\\mu M] = f_{\\text{u,p}} \\cdot (C_{\\text{max}}[\\mu M] + \\frac{F_{\\text{a}} \\cdot F_{\\text{g,perp}} \\cdot k_{\\text{a}} \\cdot Dose[mg] \\cdot 1000}{MW[g/mol] \\cdot Q_{\\text{h}} \\cdot R_{\\text{b}}})$$",
                    style="text-align: center; font-size: 1.2em; margin-top: 20px;"
                ),
                ui.tags.footer("Red if R <0.8 or R > 1.25, Green if 0.8 =< R =< 1.25"), 
                ui.tags.footer("If there are no input values for each enzyme from 'Enzyme inhibition' or 'Enzyme induction' panels, Ag, Bg, Cg, Ah, Bh or Ch are set to 1") 
            )
        )
    )
),# net effect panel end


ui.nav_panel(
    "Summary",

    ui.div(
        ui.card(
            ui.card_header("Summary", style="font-size: 24px;"),  # Card Header
            ui.card_body(
                # Action buttons
                ui.div(
                    ui.input_action_button("summary", "Create summary tables",class_="custom-summary-button")  # Summary Button
                ),
                ui.div(
                    ui.download_button("handle_download_xlsx", "Download excel file (Click after creating summary tables)",class_="custom-download-button")   # Download Button
                ),

                ui.div(
                    ui.download_button("handle_download_pdf", "Download PDF file (Click after creating summary tables)",class_="custom-download-button")   # Download Button
                ),

                # Results Card
                ui.div(
                    ui.output_table("summary_table", class_="custum-table"),  # Display results in a table
                )
 
            )#ui.card.body summary end
        )#ui.card summary end
    ),
), # summary panel end

ui.nav_panel(
    "Glossary",

    ui.div(
        ui.card(
            ui.card_header("Glossary", style="font-size: 24px;"),  # Card Header
            ui.card_body(
                ui.tags.p("Terms and Definitions", style="font-size: 20px; font-weight: bold; margin-top: 20px;"), 
                ui.output_table("glossary_table", class_="custum-table"),  
                ui.tags.p("Useful Links", style="font-size: 20px; font-weight: bold; margin-top: 20px;"), 
                ui.tags.a("M12 Drug Interaction Studies (US FDA, Aug 2024)", href="https://www.fda.gov/media/161199/download", target="_blank", style="display: block; margin-top: 10px;"),
                ui.tags.a("M12 Drug Interaction Studies: Q&A (US FDA, Aug 2024)", href="https://www.fda.gov/media/180488/download", target="_blank", style="display: block; margin-top: 10px;"),
                ui.tags.a("ICH M12 Guideline on drug interaction studies - Scientific guideline (EMA, HP link)", href="https://www.ema.europa.eu/en/ich-m12-drug-interaction-studies-scientific-guideline", target="_blank", style="display: block; margin-top: 10px;"),
                ui.tags.a("ICH M12 薬物相互作用試験 (PMDA, HP link)", href="https://www.pmda.go.jp/int-activities/int-harmony/ich/0101.html", target="_blank", style="display: block; margin-top: 10px;"),
                ui.tags.a("CDE公开征求ICH《M12：药物相互作用》指导原则及问答文件实施建议和中文版意见 (NMPA, HP link)", href="https://www.cnpharm.com/c/2024-07-18/1050390.shtml", target="_blank", style="display: block; margin-top: 10px;"),
            )#ui.card.body glossary end
        )#ui.card glossary end
    ),
), # glossary panel end

),# ui.navset_pill_list end


# Go to top button
ui.div(
    ui.input_action_button("top_button", "Go to Top"),
    style="position: fixed; bottom: 10px; right: 10px;"
),
ui.tags.script("""
    document.getElementById('top_button').onclick = function() {
        window.scrollTo({top: 0, behavior: 'smooth'});
    };
""")

)# app ui end


# Define the server logic
def server(input: Inputs, output: Outputs, session: Session):

    global results_df_ei, results_df_tdi, results_df_ed, results_df_ne
    global results_df_ti_ie, results_df_ti_hu, results_df_ti_ru, results_df_ti_re
    global results_df_ei_net, results_df_tdi_net, results_df_ed_net
    global unified_df1, unified_df2, unified_df3, unified_df4, user_memo_df
    global fuhep, fumic
    
    results_df_ei = pd.DataFrame()
    results_df_tdi = pd.DataFrame()
    results_df_ed = pd.DataFrame()
    results_df_ne = pd.DataFrame()
    results_df_ti_ie = pd.DataFrame()
    results_df_ti_hu = pd.DataFrame()
    results_df_ti_ru = pd.DataFrame()
    results_df_ti_re = pd.DataFrame()
    results_df_ei_net = pd.DataFrame()
    results_df_tdi_net = pd.DataFrame()
    results_df_ed_net = pd.DataFrame()
    unified_df1 = pd.DataFrame()
    unified_df2 = pd.DataFrame()
    unified_df3 = pd.DataFrame()
    unified_df4 = pd.DataFrame()
    user_memo_df = pd.DataFrame()

    # Reactive lists
    all_enzymes_ei_reactive = reactive.Value(all_enzymes_ei_unified.copy())
    custom_ei_next_id = reactive.Value(0)

    all_enzymes_tdi_reactive = reactive.Value(all_enzymes_tdi.copy())
    custom_tdi_next_id = reactive.Value(0)

    all_enzymes_ed_reactive = reactive.Value(all_enzymes_ed.copy())
    custom_ed_next_id = reactive.Value(0)

    all_enzymes_ne_reactive = reactive.Value(all_enzymes_ne_unified.copy())
    custom_ne_next_id = reactive.Value(0)
    
    all_transporters_ie_reactive = reactive.Value(all_transporters_ie.copy())
    custom_ti_ie_next_id = reactive.Value(0)

    all_transporters_hu_reactive = reactive.Value(all_transporters_hu.copy())
    custom_ti_hu_next_id = reactive.Value(0)

    all_transporters_ru_reactive = reactive.Value(all_transporters_ru.copy())
    custom_ti_ru_next_id = reactive.Value(0)

    all_transporters_re_reactive = reactive.Value(all_transporters_re.copy())
    custom_ti_re_next_id = reactive.Value(0)

    @render.image
    def image():
        from pathlib import Path

        dir = Path(__file__).resolve().parent

        img: ImgData = {"src": str(dir / "DDI_logo.png"), "width": "425px"}
        #img: ImgData = {"src": str(dir / "DDI_logo2.png"), "width": "425px"}
        #img: ImgData = {"src": str(dir / "ddi-app-logo2.png"), "width": "480px"}

        return img

    # Helper function to convert Cmax to μmol/L
    def get_cmax_umol():
        cmax_raw = input.cmax()
        mw = input.mw()
        if input.cmax_unit() == "ng_ml":
            # Convert ng/mL to μmol/L: (ng/mL) / MW(g/mol) * 1000
            return cmax_raw / mw
        else:
            # Already in μmol/L
            return cmax_raw
        
    # fu,mic
    @output
    @render.ui
    def input_ui_fumic():
        if input.fumic() == "input_fumic":
            return ui.input_numeric("input_value_fumic", "fu,mic - experimental value", value=0.1, step=0.01)
        elif input.fumic() == "austin_fumic":
            return ui.TagList(
                ui.input_numeric("log_p_d_fumic", "Log P/D: log P for pKa >7.4 and log D7.4 for pKa =<7.4", value=1, step=0.1),
                ui.input_numeric("c_value", "C: microsomal protein concentration (g/L)", value=0.1, step=0.01),
                ui.tags.footer("fu,mic = 1 / (1 + C * 10**(0.56 * log P/D - 1.41))")
                    )
        elif input.fumic() == "hallifax_fumic":
            return ui.TagList(
                ui.input_numeric("log_p_d_fumic", "Log P/D: log P for pKa >7.4 and log D7.4 for pKa =<7.4", value=1, step=0.1),
                ui.input_numeric("c_value", "C: microsomal protein concentration (g/L)", value=0.1, step=0.01),
                ui.tags.footer("fu,mic = 1 / (1 + C * 10**(0.072 * log P/D **2 + 0.067 * log P/D - 1.126))")
            )

    @reactive.Calc
    def calculate_fumic():
        if input.fumic() == "input_fumic":
            return float(input.input_value_fumic())
        elif input.fumic() == "austin_fumic":
            log_p_d_fumic = input.log_p_d_fumic()
            c_value = input.c_value()
            return 1 / (1 + c_value * 10**(0.56 * log_p_d_fumic - 1.41))
        elif input.fumic() == "hallifax_fumic":
            log_p_d_fumic = input.log_p_d_fumic()
            c_value = input.c_value()
            return 1 / (1 + c_value * 10**(0.072 * log_p_d_fumic**2 + 0.067 * log_p_d_fumic - 1.126))
    @output
    @render.text
    def result_fumic():
        fumic = calculate_fumic()
        return f"fu,mic: {round(fumic,3)}"

    def get_fu_for_enzyme_ei(enzyme_id):
        try:
            if input[f"use_fuinc_ei_{enzyme_id}"]():
                fuinc_val = input[f"fuinc_value_{enzyme_id}"]()
                if fuinc_val is not None and fuinc_val > 0:
                    return fuinc_val, "fu,inc"
        except:
            pass
        return calculate_fumic(), "fu,mic"

    def get_fu_for_enzyme_tdi(enzyme_id):
        try:
            if input[f"use_fuinc_tdi_{enzyme_id}"]():
                fuinc_val = input[f"fuinc_value_tdi_{enzyme_id}"]()
                if fuinc_val is not None and fuinc_val > 0:
                    return fuinc_val, "fu,inc"
        except:
            pass
        return calculate_fumic(), "fu,mic"

    # fu,hep
    @output
    @render.ui
    def input_ui_fuhep():
        if input.fuhep() == "input_fuhep":
            return ui.input_numeric("input_value_fuhep", "fu,hep - experimental value", value=0.1, step=0.01)
        elif input.fuhep() == "austin_fuhep":
            return ui.TagList(
                ui.input_numeric("log_p_d_fuhep", "Log P/D: log P for pKa >7.4 and log D7.4 for pKa =<7.4", value=1, step=0.1),
                ui.tags.footer("fu,hep = 1 / (1 + 10**(0.40 * log P/D - 1.38))")
            )
        elif input.fuhep() == "kilfold_fuhep":
            return ui.TagList(
                ui.input_numeric("log_p_d_fuhep", "Log P/D: log P for pKa >7.4 and log D7.4 for pKa =<7.4", value=1, step=0.1),
                ui.tags.footer("fu,hep = 1 / (1 + 0.625 * 10**(0.072 * log P/D**2 + 0.067 * log P/D - 1.126))")
            )

    @reactive.Calc
    def calculate_fuhep():
        if input.fuhep() == "input_fuhep":
            return float(input.input_value_fuhep())
        elif input.fuhep() == "austin_fuhep":
            log_p_d_fuhep = input.log_p_d_fuhep()
            return 1 / (1 + 10**(0.40 * log_p_d_fuhep - 1.38))
        elif input.fuhep() == "kilfold_fuhep":
            log_p_d_fuhep = input.log_p_d_fuhep()
            return 1 / (1 + 0.625 * 10**(0.072 * log_p_d_fuhep**2 + 0.067 * log_p_d_fuhep - 1.126))

    @output
    @render.text
    def result_fuhep():
        fuhep = calculate_fuhep()
        return f"fu,hep: {round(fuhep,3)}"

    def get_fu_for_enzyme_ed(enzyme_id):
        try:
            if input[f"use_fuinc_ed_{enzyme_id}"]():
                fuinc_val = input[f"fuinc_value_ed_{enzyme_id}"]()
                if fuinc_val is not None and fuinc_val > 0:
                    return fuinc_val, "fu,inc"
        except:
            pass
        return calculate_fuhep(), "fu,hep"

    # Get Kdeg,g
    def get_kdeg_g_for_enzyme_ne(enzyme_id, enzyme_name, kdeg_h):
        try:
            if enzyme_id.startswith("custom_ne_"):
                idx = int(enzyme_id.split("_")[-1])
                if input[f"custom_ne_use_kdeg_g_{idx}"]():
                    kdeg_g_input = input[f"custom_ne_kdeg_g_value_{idx}"]()
                    return kdeg_g_input if kdeg_g_input else kdeg_h
                else:
                    return kdeg_h
            else:
                if input[f"use_kdeg_g_ne_{enzyme_id}"]():
                    kdeg_g_input = input[f"kdeg_g_value_ne_{enzyme_id}"]()
                    return kdeg_g_input if kdeg_g_input else kdeg_h
                else:
                    return kdeg_h
        except:
            return kdeg_h

    # Custom enzyme/transporter handlers
    @reactive.Effect  
    @reactive.event(input.add_custom_ei_button)  
    def increment_custom_ei():  
        next_id = custom_ei_next_id()  
        new_enzyme = {  
            "attribute": "Competitive inhibition",  
            "name": f"Custom #{next_id + 1}",  
            "id": f"custom_ei_{next_id}",  
            "is_custom": True,  
            "custom_index": next_id,
            "probe": "Custom substrate",
            "intestinal": False  # Default: no intestinal calculation
        }  
        current = all_enzymes_ei_reactive()  
        all_enzymes_ei_reactive.set(current + [new_enzyme])  
        custom_ei_next_id.set(next_id + 1)

    @reactive.Effect  
    @reactive.event(input.add_custom_tdi_button)  
    def increment_custom_tdi():  
        next_id = custom_tdi_next_id()  
        new_enzyme = {  
            "attribute": "Time-dependent inhibition - hepatic",  
            "name": f"Custom #{next_id + 1}",  
            "id": f"custom_tdi_{next_id}",  
            "is_custom": True,  
            "custom_index": next_id,
            "probe": "Custom substrate",
            # "substrate": "custom_substrate",
            "kdeg_h": 0.0001  # Default kdeg,h value
        }  
        current = all_enzymes_tdi_reactive()  
        all_enzymes_tdi_reactive.set(current + [new_enzyme])  
        custom_tdi_next_id.set(next_id + 1)

    @reactive.Effect
    @reactive.event(input.add_custom_ed_button)
    def increment_custom_ed():
        next_id = custom_ed_next_id()
        new_enzyme = {
            "attribute": "Enzyme induction",
            "name": f"Custom #{next_id + 1}",
            "id": f"custom_ed_{next_id}",
            "is_custom": True,
            "custom_index": next_id
        }
        current = all_enzymes_ed_reactive()
        all_enzymes_ed_reactive.set(current + [new_enzyme])
        custom_ed_next_id.set(next_id + 1)

    @reactive.Effect  
    @reactive.event(input.add_custom_ne_button)  
    def increment_custom_ne():  
        next_id = custom_ne_next_id()  
        new_enzyme = {  
            "attribute": "Net effect",  
            "name": f"Custom #{next_id + 1}",  
            "id": f"custom_ne_{next_id}",  
            "is_custom": True,  
            "custom_index": next_id,
            "substrate": "Custom substrate",
            "fm": 0.5,
            "fg": 1.0
        }  
        current = all_enzymes_ne_reactive()  
        all_enzymes_ne_reactive.set(current + [new_enzyme])  
        custom_ne_next_id.set(next_id + 1)

    @reactive.Effect  
    @reactive.event(input.add_custom_ti_ie_button)  
    def increment_custom_ti_ie():  
        next_id = custom_ti_ie_next_id()  
        new_transporter = {  
            "attribute": "Transporter inhibition - intestinal efflux",  
            "name": f"Custom #{next_id + 1}",  
            "id": f"custom_ti_ie_{next_id}",  
            "is_custom": True,  
            "custom_index": next_id  
        }  
        current = all_transporters_ie_reactive()  
        all_transporters_ie_reactive.set(current + [new_transporter])  
        custom_ti_ie_next_id.set(next_id + 1)

    @reactive.Effect  
    @reactive.event(input.add_custom_ti_hu_button)  
    def increment_custom_ti_hu():  
        next_id = custom_ti_hu_next_id()  
        new_transporter = {  
            "attribute": "Transporter inhibition - hepatic uptake",  
            "name": f"Custom #{next_id + 1}",  
            "id": f"custom_ti_hu_{next_id}",  
            "is_custom": True,  
            "custom_index": next_id  
        }  
        current = all_transporters_hu_reactive()  
        all_transporters_hu_reactive.set(current + [new_transporter])  
        custom_ti_hu_next_id.set(next_id + 1)

    @reactive.Effect  
    @reactive.event(input.add_custom_ti_ru_button)  
    def increment_custom_ti_ru():  
        next_id = custom_ti_ru_next_id()  
        new_transporter = {  
            "attribute": "Transporter inhibition - renal uptake",  
            "name": f"Custom #{next_id + 1}",  
            "id": f"custom_ti_ru_{next_id}",  
            "is_custom": True,  
            "custom_index": next_id  
        }  
        current = all_transporters_ru_reactive()  
        all_transporters_ru_reactive.set(current + [new_transporter])  
        custom_ti_ru_next_id.set(next_id + 1)

    @reactive.Effect  
    @reactive.event(input.add_custom_ti_re_button)  
    def increment_custom_ti_re():  
        next_id = custom_ti_re_next_id()  
        new_transporter = {  
            "attribute": "Transporter inhibition - renal efflux",  
            "name": f"Custom #{next_id + 1}",  
            "id": f"custom_ti_re_{next_id}",  
            "is_custom": True,  
            "custom_index": next_id  
        }  
        current = all_transporters_re_reactive()  
        all_transporters_re_reactive.set(current + [new_transporter])  
        custom_ti_re_next_id.set(next_id + 1)


    # Remove handlers 
    for i in range(100):  
        def make_remove_handler_ei(index):  
            @reactive.Effect  
            @reactive.event(input[f"remove_custom_ei_{index}"])  
            def _():  
                current = all_enzymes_ei_reactive()  
                updated = []  
                for e in current:  
                    if e.get("custom_index") == index:  
                        e_copy = e.copy()  
                        e_copy["removed"] = True  
                        updated.append(e_copy)  
                    else:  
                        updated.append(e)  
                all_enzymes_ei_reactive.set(updated)  
        make_remove_handler_ei(i)

    for i in range(100):  
        def make_remove_handler_tdi(index):  
            @reactive.Effect  
            @reactive.event(input[f"remove_custom_tdi_{index}"])  
            def _():  
                current = all_enzymes_tdi_reactive()  
                updated = []  
                for e in current:  
                    if e.get("custom_index") == index:  
                        e_copy = e.copy()  
                        e_copy["removed"] = True  
                        updated.append(e_copy)  
                    else:  
                        updated.append(e)  
                all_enzymes_tdi_reactive.set(updated)  
        make_remove_handler_tdi(i)

    for i in range(100):
        def make_remove_handler_ed(index):
            @reactive.Effect
            @reactive.event(input[f"remove_custom_ed_{index}"])
            def _():
                current = all_enzymes_ed_reactive()
                updated = []
                for e in current:
                    if e.get("custom_index") == index:
                        e_copy = e.copy()
                        e_copy["removed"] = True
                        updated.append(e_copy)
                    else:
                        updated.append(e)
                all_enzymes_ed_reactive.set(updated)
        make_remove_handler_ed(i)

    for i in range(100):
        def make_remove_handler_ne(index):
            @reactive.Effect
            @reactive.event(input[f"remove_custom_ne_{index}"])
            def _():
                current = all_enzymes_ne_reactive()
                updated = []
                for e in current:
                    if e.get("custom_index") == index:
                        e_copy = e.copy()
                        e_copy["removed"] = True
                        updated.append(e_copy)
                    else:
                        updated.append(e)
                all_enzymes_ne_reactive.set(updated)
        make_remove_handler_ne(i)

    for i in range(100):  
        def make_remove_handler_ti_ie(index):  
            @reactive.Effect  
            @reactive.event(input[f"remove_custom_ti_ie_{index}"])  
            def _():  
                current = all_transporters_ie_reactive()  
                updated = []  
                for e in current:  
                    if e.get("custom_index") == index:  
                        e_copy = e.copy()  
                        e_copy["removed"] = True  
                        updated.append(e_copy)  
                    else:  
                        updated.append(e)  
                all_transporters_ie_reactive.set(updated)  
        make_remove_handler_ti_ie(i)

    for i in range(100):  
        def make_remove_handler_ti_hu(index):  
            @reactive.Effect  
            @reactive.event(input[f"remove_custom_ti_hu_{index}"])  
            def _():  
                current = all_transporters_hu_reactive()  
                updated = []  
                for e in current:  
                    if e.get("custom_index") == index:  
                        e_copy = e.copy()  
                        e_copy["removed"] = True  
                        updated.append(e_copy)  
                    else:  
                        updated.append(e)  
                all_transporters_hu_reactive.set(updated)  
        make_remove_handler_ti_hu(i)

    for i in range(100):  
        def make_remove_handler_ti_ru(index):  
            @reactive.Effect  
            @reactive.event(input[f"remove_custom_ti_ru_{index}"])  
            def _():  
                current = all_transporters_ru_reactive()  
                updated = []  
                for e in current:  
                    if e.get("custom_index") == index:  
                        e_copy = e.copy()  
                        e_copy["removed"] = True  
                        updated.append(e_copy)  
                    else:  
                        updated.append(e)  
                all_transporters_ru_reactive.set(updated)  
        make_remove_handler_ti_ru(i)

    for i in range(100):  
        def make_remove_handler_ti_re(index):  
            @reactive.Effect  
            @reactive.event(input[f"remove_custom_ti_re_{index}"])  
            def _():  
                current = all_transporters_re_reactive()  
                updated = []  
                for e in current:  
                    if e.get("custom_index") == index:  
                        e_copy = e.copy()  
                        e_copy["removed"] = True  
                        updated.append(e_copy)  
                    else:  
                        updated.append(e)  
                all_transporters_re_reactive.set(updated)  
        make_remove_handler_ti_re(i)


    # Render all custom enzyme input fields dynamically
    @output  
    @render.ui  
    def custom_ei_inputs_container():  
        enzymes = all_enzymes_ei_reactive()  
        custom_enzymes = [e for e in enzymes if e.get("is_custom", False) and not e.get("removed", False)]  
        
        if not custom_enzymes:  
            return ui.div()  
        
        input_fields = []  
        for enzyme in custom_enzymes:  
            i = enzyme["custom_index"]  
            input_fields.append(  
                ui.div(  
                    ui.tags.hr(style="border-top: 2px dashed #17a2b8;"),  
                    ui.tags.h6(  
                        f"Custom Enzyme #{i+1}",  
                        ui.input_action_button(  
                            f"remove_custom_ei_{i}",  
                            "✕ Remove",  
                            class_="btn btn-danger btn-sm",  
                            style="float: right; font-size: 10px; padding: 2px 8px;"  
                        ),  
                        style="color: #0c5460; margin-bottom: 10px;"  
                    ),  
                    ui.input_text(f"custom_ei_name_{i}", "Enzyme name:", value="", placeholder="e.g., CYP2E1"),  
                    ui.input_text(f"custom_ei_probe_{i}", "Probe substrate:", value="", placeholder="e.g., Midazolam"), 
                    ui.input_numeric(f"custom_ei_ki_{i}", "Ki (μmol/L):", value=10, step=0.1), 
                    ui.input_checkbox(f"custom_ei_intestinal_{i}", "Include intestinal calculation", value=False),  
                    ui.input_checkbox(f"use_fuinc_custom_ei_{i}", "Use individual fu,inc instead of fu,mic", value=False),  
                    ui.output_ui(f"fuinc_input_custom_ei_{i}"),  
                    style="margin-left: 20px; padding: 15px; background-color: #d1ecf1; border: 2px solid #17a2b8; border-radius: 5px; margin-bottom: 10px;"  
                )  
            )  
        return ui.div(*input_fields)

    @output  
    @render.ui  
    def custom_tdi_inputs_container():  
        enzymes = all_enzymes_tdi_reactive()  
        custom_enzymes = [e for e in enzymes if e.get("is_custom", False) and not e.get("removed", False)]  
        
        if not custom_enzymes:  
            return ui.div()  
        
        input_fields = []  
        for enzyme in custom_enzymes:  
            i = enzyme["custom_index"]  
            input_fields.append(  
                ui.div(  
                    ui.tags.hr(style="border-top: 2px dashed #dc3545;"),  
                    ui.tags.h6(  
                        f"Custom Enzyme #{i+1}",  
                        ui.input_action_button(  
                            f"remove_custom_tdi_{i}",  
                            "✕ Remove",  
                            class_="btn btn-danger btn-sm",  
                            style="float: right; font-size: 10px; padding: 2px 8px;"  
                        ),  
                        style="color: #721c24; margin-bottom: 10px;"  
                    ),  
                    ui.input_text(f"custom_tdi_name_{i}", "Enzyme name:", value="", placeholder="e.g., CYP3A4"),  
                    ui.input_text(f"custom_tdi_probe_{i}", "Probe substrate:", value="", placeholder="e.g., Midazolam"),  
                    ui.input_numeric(f"custom_tdi_ki_tdi_{i}", "KI (μmol/L):", value=10, step=0.1),
                    ui.input_numeric(f"custom_tdi_kinact_{i}", "Kinact (/h):", value=0.1, step=0.01),
                    ui.input_numeric(f"custom_tdi_kdeg_h_{i}", "Kdeg,h (/h):", value=0.0193, step=0.0001, min=0),  
                    ui.input_checkbox(f"use_fuinc_custom_tdi_{i}", "Use individual fu,inc instead of fu,mic", value=False),  
                    ui.output_ui(f"fuinc_input_custom_tdi_{i}"),  
                    style="margin-left: 20px; padding: 15px; background-color: #f8d7da; border: 2px solid #dc3545; border-radius: 5px; margin-bottom: 10px;"  
                )  
            )  
        return ui.div(*input_fields)

    @output
    @render.ui
    def custom_ed_inputs_container():
        enzymes = all_enzymes_ed_reactive()
        custom_enzymes = [e for e in enzymes if e.get("is_custom", False) and not e.get("removed", False)]
        
        if not custom_enzymes:
            return ui.div()
        
        input_fields = []
        for enzyme in custom_enzymes:
            i = enzyme["custom_index"]
            input_fields.append(
                ui.div(
                    ui.tags.hr(style="border-top: 2px dashed #ffc107;"),
                    ui.tags.h6(
                        f"Custom Enzyme #{i+1}",
                        ui.input_action_button(
                            f"remove_custom_ed_{i}",
                            "✕ Remove",
                            class_="btn btn-danger btn-sm",
                            style="float: right; font-size: 10px; padding: 2px 8px;"
                        ),
                        style="color: #856404; margin-bottom: 10px;"
                    ),
                    ui.input_text(f"custom_ed_name_{i}", "Enzyme name:", value="", placeholder="e.g., CYP2E1"),
                    ui.input_numeric(f"custom_ed_emax_{i}", "Emax:", value=0.3, step=0.1, min=0),
                    ui.input_numeric(f"custom_ed_ec50_{i}", "EC50 (μmol/L):", value=20, step=1, min=0),
                    ui.input_checkbox(f"use_fuinc_custom_ed_{i}", "Use individual fu,inc instead of fu,hep", value=False),
                    ui.output_ui(f"fuinc_input_custom_ed_{i}"),
                    style="margin-left: 20px; padding: 15px; background-color: #fff3cd; border: 2px solid #ffc107; border-radius: 5px; margin-bottom: 10px;"
                )
            )
        return ui.div(*input_fields)
    
    @output
    @render.ui
    def custom_ne_inputs_container():
        enzymes = all_enzymes_ne_reactive()
        custom_enzymes = [e for e in enzymes if e.get("is_custom", False) and not e.get("removed", False)]
        
        if not custom_enzymes:
            return ui.div()
        
        input_fields = []
        for enzyme in custom_enzymes:
            i = enzyme["custom_index"]
            input_fields.append(
                ui.div(
                    ui.tags.hr(style="border-top: 2px dashed #17a2b8;"),
                    ui.tags.h6(
                        f"Custom Enzyme #{i+1}",
                        ui.input_action_button(
                            f"remove_custom_ne_{i}",
                            "✕ Remove",
                            class_="btn btn-danger btn-sm",
                            style="float: right; font-size: 10px; padding: 2px 8px;"
                        ),
                        style="color: #0c5460; margin-bottom: 10px;"
                    ),
                    ui.input_text(f"custom_ne_name_{i}", "Enzyme name:", value="", placeholder="e.g., CYP2C9"),
                    ui.input_text(f"custom_ne_substrate_{i}", "Clinical substrate name:", value="", placeholder="e.g., Warfarin"),
                    ui.input_numeric(f"custom_ne_fm_{i}", "fm (fraction metabolized):", value=0.5, step=0.01, min=0, max=1),
                    ui.input_numeric(f"custom_ne_fg_{i}", "Fg,victim (intestinal availability):", value=1.0, step=0.01, min=0, max=1),
                    ui.input_checkbox(f"custom_ne_use_kdeg_g_{i}", "Use different Kdeg,g (intestinal) instead of Kdeg,h", value=False),
                    ui.output_ui(f"custom_ne_kdeg_g_input_{i}"),
                    style="margin-left: 20px; padding: 15px; background-color: #d1ecf1; border: 2px solid #17a2b8; border-radius: 5px; margin-bottom: 10px;"
                )
            )
        return ui.div(*input_fields)

    @output  
    @render.ui  
    def custom_ti_ie_inputs_container():  
        transporters = all_transporters_ie_reactive()  
        custom_transporters = [t for t in transporters if t.get("is_custom", False) and not t.get("removed", False)]  
        
        if not custom_transporters:  
            return ui.div()  
        
        input_fields = []  
        for transporter in custom_transporters:  
            i = transporter["custom_index"]  
            input_fields.append(  
                ui.div(  
                    ui.tags.hr(style="border-top: 2px dashed #28a745;"),  
                    ui.tags.h6(  
                        f"Custom Transporter #{i+1}",  
                        ui.input_action_button(  
                            f"remove_custom_ti_ie_{i}",  
                            "✕ Remove",  
                            class_="btn btn-danger btn-sm",  
                            style="float: right; font-size: 10px; padding: 2px 8px;"  
                        ),  
                        style="color: #155724; margin-bottom: 10px;"  
                    ),  
                    ui.input_text(f"custom_ti_ie_name_{i}", "Transporter name:", value="", placeholder="e.g., P-gp"),  
                    ui.input_numeric(f"custom_ti_ie_ic50_{i}", "IC50 or Ki (μmol/L):", value=10, step=0.1),
                    ui.input_checkbox(f"use_fuinc_custom_ti_ie_{i}", "Use individual fu,inc instead of shared fu,inc", value=False),  
                    ui.output_ui(f"fuinc_input_custom_ti_ie_{i}"),  
                    style="margin-left: 20px; padding: 15px; background-color: #d4edda; border: 2px solid #28a745; border-radius: 5px; margin-bottom: 10px;"  
                )  
            )  
        return ui.div(*input_fields)

    @output  
    @render.ui  
    def custom_ti_hu_inputs_container():  
        transporters = all_transporters_hu_reactive()  
        custom_transporters = [t for t in transporters if t.get("is_custom", False) and not t.get("removed", False)]  
        
        if not custom_transporters:  
            return ui.div()  
        
        input_fields = []  
        for transporter in custom_transporters:  
            i = transporter["custom_index"]  
            input_fields.append(  
                ui.div(  
                    ui.tags.hr(style="border-top: 2px dashed #6f42c1;"),  
                    ui.tags.h6(  
                        f"Custom Transporter #{i+1}",  
                        ui.input_action_button(  
                            f"remove_custom_ti_hu_{i}",  
                            "✕ Remove",  
                            class_="btn btn-danger btn-sm",  
                            style="float: right; font-size: 10px; padding: 2px 8px;"  
                        ),  
                        style="color: #3d1a78; margin-bottom: 10px;"  
                    ),  
                    ui.input_text(f"custom_ti_hu_name_{i}", "Transporter name:", value="", placeholder="e.g., OATP1B1"),  
                    ui.input_numeric(f"custom_ti_hu_ic50_{i}", "IC50 or Ki (μmol/L):", value=10, step=0.1),
                    ui.input_checkbox(f"use_fuinc_custom_ti_hu_{i}", "Use individual fu,inc instead of shared fu,inc", value=False),  
                    ui.output_ui(f"fuinc_input_custom_ti_hu_{i}"),  
                    style="margin-left: 20px; padding: 15px; background-color: #e2d9f3; border: 2px solid #6f42c1; border-radius: 5px; margin-bottom: 10px;"  
                )  
            )  
        return ui.div(*input_fields)

    @output  
    @render.ui  
    def custom_ti_ru_inputs_container():  
        transporters = all_transporters_ru_reactive()  
        custom_transporters = [t for t in transporters if t.get("is_custom", False) and not t.get("removed", False)]  
        
        if not custom_transporters:  
            return ui.div()  
        
        input_fields = []  
        for transporter in custom_transporters:  
            i = transporter["custom_index"]  
            input_fields.append(  
                ui.div(  
                    ui.tags.hr(style="border-top: 2px dashed #fd7e14;"),  
                    ui.tags.h6(  
                        f"Custom Transporter #{i+1}",  
                        ui.input_action_button(  
                            f"remove_custom_ti_ru_{i}",  
                            "✕ Remove",  
                            class_="btn btn-danger btn-sm",  
                            style="float: right; font-size: 10px; padding: 2px 8px;"  
                        ),  
                        style="color: #8a4a00; margin-bottom: 10px;"  
                    ),  
                    ui.input_text(f"custom_ti_ru_name_{i}", "Transporter name:", value="", placeholder="e.g., OAT1"), 
                    ui.input_numeric(f"custom_ti_ru_ic50_{i}", "IC50 or Ki (μmol/L):", value=10, step=0.1), 
                    ui.input_checkbox(f"use_fuinc_custom_ti_ru_{i}", "Use individual fu,inc instead of shared fu,inc", value=False),  
                    ui.output_ui(f"fuinc_input_custom_ti_ru_{i}"),  
                    style="margin-left: 20px; padding: 15px; background-color: #ffeaa7; border: 2px solid #fd7e14; border-radius: 5px; margin-bottom: 10px;"  
                )  
            )  
        return ui.div(*input_fields)

    @output  
    @render.ui  
    def custom_ti_re_inputs_container():  
        transporters = all_transporters_re_reactive()  
        custom_transporters = [t for t in transporters if t.get("is_custom", False) and not t.get("removed", False)]  
        
        if not custom_transporters:  
            return ui.div()  
        
        input_fields = []  
        for transporter in custom_transporters:  
            i = transporter["custom_index"]  
            input_fields.append(  
                ui.div(  
                    ui.tags.hr(style="border-top: 2px dashed #6c757d;"),  
                    ui.tags.h6(  
                        f"Custom Transporter #{i+1}",  
                        ui.input_action_button(  
                            f"remove_custom_ti_re_{i}",  
                            "✕ Remove",  
                            class_="btn btn-danger btn-sm",  
                            style="float: right; font-size: 10px; padding: 2px 8px;"  
                        ),  
                        style="color: #383d41; margin-bottom: 10px;"  
                    ),  
                    ui.input_text(f"custom_ti_re_name_{i}", "Transporter name:", value="", placeholder="e.g., MRP2"), 
                    ui.input_numeric(f"custom_ti_re_ic50_{i}", "IC50 or Ki (μmol/L):", value=10, step=0.1), 
                    ui.input_checkbox(f"use_fuinc_custom_ti_re_{i}", "Use individual fu,inc instead of shared fu,inc", value=False),  
                    ui.output_ui(f"fuinc_input_custom_ti_re_{i}"),  
                    style="margin-left: 20px; padding: 15px; background-color: #ffeaa7; border: 2px solid #fd7e14; border-radius: 5px; margin-bottom: 10px;"  
                )  
            )  
        return ui.div(*input_fields)


    # Enzyme/transporter selection UI
    @output  
    @render.ui  
    def enzyme_selection_ui_ei():  
        enzymes = all_enzymes_ei_reactive()  
        visible_enzymes = [e for e in enzymes if not e.get("removed", False)]  
        return ui.div(  
            *[  
                ui.div(  
                    ui.input_checkbox(  
                        f'select_{enzyme["id"]}',  
                        enzyme["name"] if enzyme["name"] else f'Custom #{enzyme.get("custom_index", 0)+1}',  
                        value=enzyme.get("is_custom", False)   
                    ),  
                    ui.output_ui(f'parameters_{enzyme["id"]}'),  
                    class_="enzyme-block"  
                )  
                for enzyme in visible_enzymes  
            ]  
        )

    @output  
    @render.ui  
    def enzyme_selection_ui_tdi():  
        enzymes = all_enzymes_tdi_reactive()  
        visible_enzymes = [e for e in enzymes if not e.get("removed", False)]  
        return ui.div(  
            *[  
                ui.div(  
                    ui.input_checkbox(  
                        f'select_{enzyme["id"]}',  
                        enzyme["name"] if enzyme["name"] else f'Custom #{enzyme.get("custom_index", 0)+1}',  
                        value=enzyme.get("is_custom", False) 
                    ),  
                    ui.output_ui(f'parameters_{enzyme["id"]}'),  
                    class_="enzyme-block"  
                )  
                for enzyme in visible_enzymes  
            ]  
        )

    @output  
    @render.ui  
    def enzyme_selection_ui_ed():  
        enzymes = all_enzymes_ed_reactive()  
        visible_enzymes = [e for e in enzymes if not e.get("removed", False)]  
        return ui.div(  
            *[  
                ui.div(  
                    ui.input_checkbox(  
                        f'select_{enzyme["id"]}',  
                        enzyme["name"] if enzyme["name"] else f'Custom #{enzyme.get("custom_index", 0)+1}',  
                        value=enzyme.get("is_custom", False)  
                    ),  
                    ui.output_ui(f'parameters_{enzyme["id"]}'),  
                    class_="enzyme-block"  
                )  
                for enzyme in visible_enzymes  
            ]  
        )

    @output  
    @render.ui  
    def enzyme_selection_ui_ne():  
        enzymes = all_enzymes_ne_reactive()  
        visible_enzymes = [e for e in enzymes if not e.get("removed", False)]  
        return ui.div(  
            *[  
                ui.div(  
                    ui.input_checkbox(  
                        f'select_{enzyme["id"]}',  
                        enzyme["name"] if enzyme["name"] else f'Custom #{enzyme.get("custom_index", 0)+1}',  
                        value=enzyme.get("is_custom", False)  
                    ),  
                    ui.output_ui(f'parameters_{enzyme["id"]}'),  
                    class_="enzyme-block"  
                )  
                for enzyme in visible_enzymes  
            ]  
        )

    @output  
    @render.ui  
    def transporter_selection_ui_ti_ie():  
        transporters = all_transporters_ie_reactive()  
        visible_transporters = [t for t in transporters if not t.get("removed", False)]  
        return ui.div(  
            *[  
                ui.div(  
                    ui.input_checkbox(  
                        f'select_{transporter["id"]}',  
                        transporter["name"] if transporter["name"] else f'Custom #{transporter.get("custom_index", 0)+1}',  
                        value=transporter.get("is_custom", False) 
                    ),  
                    ui.output_ui(f'parameters_{transporter["id"]}'),  
                    class_="transporter-block"  
                )  
                for transporter in visible_transporters  
            ]  
        )

    @output  
    @render.ui  
    def transporter_selection_ui_ti_hu():  
        transporters = all_transporters_hu_reactive()  
        visible_transporters = [t for t in transporters if not t.get("removed", False)]  
        return ui.div(  
            *[  
                ui.div(  
                    ui.input_checkbox(  
                        f'select_{transporter["id"]}',  
                        transporter["name"] if transporter["name"] else f'Custom #{transporter.get("custom_index", 0)+1}',  
                        value=transporter.get("is_custom", False)  
                    ),  
                    ui.output_ui(f'parameters_{transporter["id"]}'),  
                    class_="transporter-block"  
                )  
                for transporter in visible_transporters  
            ]  
        )

    @output  
    @render.ui  
    def transporter_selection_ui_ti_ru():  
        transporters = all_transporters_ru_reactive()  
        visible_transporters = [t for t in transporters if not t.get("removed", False)]  
        return ui.div(  
            *[  
                ui.div(  
                    ui.input_checkbox(  
                        f'select_{transporter["id"]}',  
                        transporter["name"] if transporter["name"] else f'Custom #{transporter.get("custom_index", 0)+1}',  
                        value=transporter.get("is_custom", False)   
                    ),  
                    ui.output_ui(f'parameters_{transporter["id"]}'),  
                    class_="transporter-block"  
                )  
                for transporter in visible_transporters  
            ]  
        )

    @output  
    @render.ui  
    def transporter_selection_ui_ti_re():  
        transporters = all_transporters_re_reactive()  
        visible_transporters = [t for t in transporters if not t.get("removed", False)]  
        return ui.div(  
            *[  
                ui.div(  
                    ui.input_checkbox(  
                        f'select_{transporter["id"]}',  
                        transporter["name"] if transporter["name"] else f'Custom #{transporter.get("custom_index", 0)+1}',  
                        value=transporter.get("is_custom", False) 
                    ),  
                    ui.output_ui(f'parameters_{transporter["id"]}'),  
                    class_="transporter-block"  
                )  
                for transporter in visible_transporters  
            ]  
        )


    # Dynamically render parameters for each enzyme and transporter 
    def make_render_parameters_ei_unified(enzyme):
        @output(id=f"parameters_{enzyme['id']}")
        @render.ui
        def render_parameters():
            if input[f"select_{enzyme['id']}"]():
                base_ui = ui.div(
                    ui.input_text(f"probe_{enzyme['id']}", f"Probe substrate for {enzyme['name']}:", value=enzyme.get("probe", "N/A")),
                    ui.input_select(f"ki_input_type_{enzyme['id']}", "Input type:", {"ki_direct": "Ki (direct input)", "ic50_convert": "IC50 (convert to Ki using Cheng-Prusoff)"}),
                    ui.output_ui(f"ki_input_ui_{enzyme['id']}"),
                    ui.output_text_verbatim(f"ki_result_{enzyme['id']}"),
                    ui.input_checkbox(f"use_fuinc_ei_{enzyme['id']}",f"Use fu,inc for {enzyme['name']} instead of fu,mic", value=False),
                    ui.output_ui(f"fuinc_input_ei_{enzyme['id']}"),
                )

                # For CYP3A4: show note that intestinal calculation will also be performed
                if enzyme.get("intestinal", False):
                    return ui.div(
                        base_ui,
                        ui.tags.p(
                            "※ CYP3A4: The same Ki will be used for both hepatic and intestinal calculations.",
                            style="font-size: 12px; color: #4B0082; font-style: italic; margin-top: 5px;"
                        )
                    )
                return base_ui
            else:
                return None    

        @output(id=f"fuinc_input_ei_{enzyme['id']}")
        @render.ui
        def fuinc_input_ei_ui():
            if input[f"select_{enzyme['id']}"]():
                try:
                    use_fuinc = input[f"use_fuinc_ei_{enzyme['id']}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_{enzyme['id']}",
                            f"fu,inc for {enzyme['name']} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
            return ui.div()

    def make_render_parameters_tdi(enzyme):
        @output(id=f"parameters_{enzyme['id']}")
        @render.ui
        def render_parameters():
            if input[f"select_{enzyme['id']}"]():
                return ui.div(
                    ui.input_text(f"probe_tdi_{enzyme['id']}", f"Probe substrate for {enzyme['name']}:", value=enzyme.get("probe", "N/A")),
                    ui.input_numeric(f"ki_tdi_{enzyme['id']}", f"KI ({enzyme['name']}_{enzyme['substrate']}) (μmol/L)", value=10, step=0.1),
                    ui.input_numeric(f"kinact_{enzyme['id']}", f"Kinact ({enzyme['name']}_{enzyme['substrate']}) (/h)", value=0.1, step=0.01),                   
                    ui.input_numeric(f"kdeg_h_{enzyme['id']}", f"Kdeg,h ({enzyme['name']}_{enzyme['substrate']}) (/h)", value=enzyme['kdeg_h'], step=0.0001), 
                    ui.input_checkbox(f"use_fuinc_tdi_{enzyme['id']}", f"Use fu,inc for {enzyme['name']} instead of fu,mic", value=False),
                    ui.output_ui(f"fuinc_input_tdi_{enzyme['id']}"),                  
                )
            else:
                return None

        @output(id=f"fuinc_input_tdi_{enzyme['id']}")
        @render.ui
        def fuinc_input_tdi_ui():
            if input[f"select_{enzyme['id']}"]():
                try:
                    use_fuinc = input[f"use_fuinc_tdi_{enzyme['id']}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_tdi_{enzyme['id']}",
                            f"fu,inc for {enzyme['name']} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
            return ui.div()

    def make_render_parameters_ed(enzyme):
        @output(id=f"parameters_{enzyme['id']}")
        @render.ui
        def render_parameters():
            if input[f"select_{enzyme['id']}"]():
                return ui.div(
                    ui.input_numeric(f"emax_{enzyme['id']}", f"Emax ({enzyme['name']})", value=0.3, step=0.1),
                    ui.input_numeric(f"ec50_{enzyme['id']}", f"EC50 ({enzyme['name']}) (μmol/L)", value=20, step=1),
                    ui.input_checkbox(f"use_fuinc_ed_{enzyme['id']}", f"Use fu,inc for {enzyme['name']} instead of fu,hep", value=False),
                    ui.output_ui(f"fuinc_input_ed_{enzyme['id']}"), 
                )
            else:
                return None

        @output(id=f"fuinc_input_ed_{enzyme['id']}")
        @render.ui
        def fuinc_input_ed_ui():
            if input[f"select_{enzyme['id']}"]():
                try:
                    use_fuinc = input[f"use_fuinc_ed_{enzyme['id']}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_ed_{enzyme['id']}",
                            f"fu,inc for {enzyme['name']} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
            return ui.div()

    def make_render_parameters_ne_unified(enzyme):
        defaults = {"fm": enzyme.get("fm", 0.5), "fg": enzyme.get("fg", 1.0)}
        @output(id=f"parameters_{enzyme['id']}")
        @render.ui
        def render_parameters():
            if input[f"select_{enzyme['id']}"]():
                return ui.div(
                    ui.tags.h6(f"Clinical substrate parameters for {enzyme['name']}"),
                    ui.tags.p(
                        f"Default clinical substrate: {enzyme.get('substrate', 'N/A')}",
                        style="font-size: 12px; color: gray;"
                    ),

                    # --- Default substrate block (editable) ---
                    ui.div(
                        # Editable substrate name (default substrate)
                        ui.input_text(
                            f"substrate_name_{enzyme['id']}",
                            "Clinical substrate name (editable):",
                            value=enzyme.get('substrate', 'Custom substrate')
                        ),

                        # fm input for default substrate
                        ui.input_numeric(
                            f"fm_{enzyme['id']}",
                            f"fm (fraction metabolized by {enzyme['name']})",
                            value=defaults["fm"], step=0.01, max=1, min=0
                        ),

                        # Fg,victim input for default substrate
                        ui.input_numeric(
                            f"fg_victim_{enzyme['id']}",
                            "Fg,victim (intestinal availability of victim drug)",
                            value=defaults["fg"], step=0.01, max=1, min=0
                        ),

                        # Kdeg,g input (if applicable)
                        ui.input_checkbox(
                            f"use_kdeg_g_ne_{enzyme['id']}", 
                            "Use different Kdeg,g (intestinal) instead of Kdeg,h (hepatic)", 
                            value=False
                        ),
                        ui.output_ui(f"kdeg_g_input_ne_{enzyme['id']}"),

                        style="padding: 10px; background-color: #f0f8ff; border-radius: 5px; margin-bottom: 10px;"
                    ),
                    ui.tags.small(
                        "You can edit the default substrate name and parameters",
                        style="color: #666;"
                    )
                )
            else:
                return None
            
        @output(id=f"kdeg_g_input_ne_{enzyme['id']}")
        @render.ui
        def kdeg_g_input_ne_ui():
            if input[f"select_{enzyme['id']}"]():
                try:
                    use_kdeg_g = input[f"use_kdeg_g_ne_{enzyme['id']}"]()
                except:
                    use_kdeg_g = False
                if use_kdeg_g:
                    enzyme_name = enzyme["name"]
                    default_kdeg_g = next(
                        (e["kdeg_h"] for e in all_enzymes_tdi if e["name"] == enzyme_name),
                        0.0001  
                    )
                    
                    return ui.input_numeric(
                        f"kdeg_g_value_ne_{enzyme['id']}",
                        "Kdeg,g (intestinal) (/h):",
                        value=default_kdeg_g,
                        step=0.0001,
                        min=0
                    )
            return ui.div()

    def make_render_parameters_ti_ie(transporter):
        @output(id=f"parameters_{transporter['id']}")
        @render.ui
        def render_parameters():
            if input[f"select_{transporter['id']}"]():
                return ui.div(
                    ui.input_numeric(f"ic50_{transporter['id']}", f"IC50 or Ki ({transporter['name']}) (μmol/L)", value=10, step=0.1),
                    ui.input_checkbox(
                        f"use_fuinc_ti_{transporter['id']}",
                        f"Use individual fu,inc for {transporter['name']} instead of shared fu,inc",
                        value=False
                    ),
                    ui.output_ui(f"fuinc_input_ti_{transporter['id']}"),
                )
            else:
                return None

        @output(id=f"fuinc_input_ti_{transporter['id']}")
        @render.ui
        def fuinc_input_ti_ui():
            if input[f"select_{transporter['id']}"]():
                try:
                    use_fuinc = input[f"use_fuinc_ti_{transporter['id']}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_ti_{transporter['id']}",
                            f"fu,inc for {transporter['name']} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
            return ui.div()

    def make_render_parameters_ti_hu(transporter):
        @output(id=f"parameters_{transporter['id']}")
        @render.ui
        def render_parameters():
            if input[f"select_{transporter['id']}"]():
                return ui.div(
                    ui.input_numeric(f"ic50_{transporter['id']}", f"IC50 or Ki ({transporter['name']}) (μmol/L)", value=10, step=0.1),
                    ui.input_checkbox(
                        f"use_fuinc_ti_{transporter['id']}",
                        f"Use individual fu,inc for {transporter['name']} instead of shared fu,inc",
                        value=False
                    ),
                    ui.output_ui(f"fuinc_input_ti_{transporter['id']}"),
                )
            else:
                return None

        @output(id=f"fuinc_input_ti_{transporter['id']}")
        @render.ui
        def fuinc_input_ti_ui():
            if input[f"select_{transporter['id']}"]():
                try:
                    use_fuinc = input[f"use_fuinc_ti_{transporter['id']}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_ti_{transporter['id']}",
                            f"fu,inc for {transporter['name']} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
            return ui.div()
        
    def make_render_parameters_ti_ru(transporter):
        @output(id=f"parameters_{transporter['id']}")
        @render.ui
        def render_parameters():
            if input[f"select_{transporter['id']}"]():
                return ui.div(
                    ui.input_numeric(f"ic50_{transporter['id']}", f"IC50 or Ki ({transporter['name']}) (μmol/L)", value=10, step=0.1),
                    ui.input_checkbox(
                        f"use_fuinc_ti_{transporter['id']}",
                        f"Use individual fu,inc for {transporter['name']} instead of shared fu,inc",
                        value=False
                    ),
                    ui.output_ui(f"fuinc_input_ti_{transporter['id']}"),
                )
            else:
                return None

        @output(id=f"fuinc_input_ti_{transporter['id']}")
        @render.ui
        def fuinc_input_ti_ui():
            if input[f"select_{transporter['id']}"]():
                try:
                    use_fuinc = input[f"use_fuinc_ti_{transporter['id']}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_ti_{transporter['id']}",
                            f"fu,inc for {transporter['name']} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
            return ui.div()
        
    def make_render_parameters_ti_re(transporter):
        @output(id=f"parameters_{transporter['id']}")
        @render.ui
        def render_parameters():
            if input[f"select_{transporter['id']}"]():
                return ui.div(
                    ui.input_numeric(f"ic50_{transporter['id']}", f"IC50 or Ki ({transporter['name']}) (μmol/L)", value=10, step=0.1),
                    ui.input_checkbox(
                        f"use_fuinc_ti_{transporter['id']}",
                        f"Use individual fu,inc for {transporter['name']} instead of shared fu,inc",
                        value=False
                    ),
                    ui.output_ui(f"fuinc_input_ti_{transporter['id']}"),
                )
            else:
                return None

        @output(id=f"fuinc_input_ti_{transporter['id']}")
        @render.ui
        def fuinc_input_ti_ui():
            if input[f"select_{transporter['id']}"]():
                try:
                    use_fuinc = input[f"use_fuinc_ti_{transporter['id']}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_ti_{transporter['id']}",
                            f"fu,inc for {transporter['name']} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
            return ui.div()

    # Custom fu,inc input
    for i in range(100):
        def make_custom_fuinc_input_ei(index):
            @output(id=f"fuinc_input_custom_ei_{index}")
            @render.ui
            def fuinc_input_custom_ei_ui():
                try:
                    use_fuinc = input[f"use_fuinc_custom_ei_{index}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_custom_ei_{index}",
                            f"fu,inc for Custom Enzyme #{index+1} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
                return ui.div()
        make_custom_fuinc_input_ei(i)

    for i in range(100):
        def make_custom_fuinc_input_tdi(index):
            @output(id=f"fuinc_input_custom_tdi_{index}")
            @render.ui
            def fuinc_input_custom_tdi_ui():
                try:
                    use_fuinc = input[f"use_fuinc_custom_tdi_{index}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_custom_tdi_{index}",
                            f"fu,inc for Custom Enzyme #{index+1} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
                return ui.div()
        make_custom_fuinc_input_tdi(i)

    for i in range(100):
        def make_custom_fuinc_input_ed(index):
            @output(id=f"fuinc_input_custom_ed_{index}")
            @render.ui
            def fuinc_input_custom_ed_ui():
                try:
                    use_fuinc = input[f"use_fuinc_custom_ed_{index}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_custom_ed_{index}",
                            f"fu,inc for Custom Enzyme #{index+1} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
                return ui.div()
        make_custom_fuinc_input_ed(i)

    for i in range(100):
        def make_custom_fuinc_input_ti_ie(index):
            @output(id=f"fuinc_input_custom_ti_ie_{index}")
            @render.ui
            def fuinc_input_custom_ti_ie_ui():
                try:
                    use_fuinc = input[f"use_fuinc_custom_ti_ie_{index}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_custom_ti_ie_{index}",
                            f"fu,inc for Custom Transporter #{index+1} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
                return ui.div()
        make_custom_fuinc_input_ti_ie(i)

    for i in range(100):
        def make_custom_fuinc_input_ti_hu(index):
            @output(id=f"fuinc_input_custom_ti_hu_{index}")
            @render.ui
            def fuinc_input_custom_ti_hu_ui():
                try:
                    use_fuinc = input[f"use_fuinc_custom_ti_hu_{index}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_custom_ti_hu_{index}",
                            f"fu,inc for Custom Transporter #{index+1} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
                return ui.div()
        make_custom_fuinc_input_ti_hu(i)

    for i in range(100):
        def make_custom_fuinc_input_ti_ru(index):
            @output(id=f"fuinc_input_custom_ti_ru_{index}")
            @render.ui
            def fuinc_input_custom_ti_ru_ui():
                try:
                    use_fuinc = input[f"use_fuinc_custom_ti_ru_{index}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_custom_ti_ru_{index}",
                            f"fu,inc for Custom Transporter #{index+1} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
                return ui.div()
        make_custom_fuinc_input_ti_ru(i)

    for i in range(100):
        def make_custom_fuinc_input_ti_re(index):
            @output(id=f"fuinc_input_custom_ti_re_{index}")
            @render.ui
            def fuinc_input_custom_ti_re_ui():
                try:
                    use_fuinc = input[f"use_fuinc_custom_ti_re_{index}"]()
                except:
                    use_fuinc = False
                if use_fuinc:
                    return ui.div(
                        ui.input_numeric(
                            f"fuinc_value_custom_ti_re_{index}",
                            f"fu,inc for Custom Transporter #{index+1} (experimental value)",
                            value=0.1, step=0.01, min=0, max=1
                        ),
                        style="margin-left: 20px; padding: 10px; background-color: #ffe6cc; border-radius: 5px;"
                    )
                return ui.div()
        make_custom_fuinc_input_ti_re(i)

    # Custom kdeg_g input
    for i in range(100):
        def make_custom_kdeg_g_input_ne(index):
            @output(id=f"custom_ne_kdeg_g_input_{index}")
            @render.ui
            def custom_kdeg_g_input_ne_ui():
                try:
                    use_kdeg_g = input[f"custom_ne_use_kdeg_g_{index}"]()
                except:
                    use_kdeg_g = False
                if use_kdeg_g:
                    return ui.input_numeric(
                        f"custom_ne_kdeg_g_value_{index}",
                        "Kdeg,g (intestinal) (/h):",
                        value=0.0193,
                        step=0.0001,
                        min=0
                    )
                return ui.div()
        make_custom_kdeg_g_input_ne(i)

    # Create dynamic UI for Ki input (enzyme inhibition)
    def make_ki_input_ui(enzyme):
        @output(id=f"ki_input_ui_{enzyme['id']}")
        @render.ui
        def ki_input_ui():
            if input[f"select_{enzyme['id']}"]():
                if input[f"ki_input_type_{enzyme['id']}"]() == "ki_direct":
                    # Label shows enzyme name only, no substrate name
                    return ui.input_numeric(f"ki_value_{enzyme['id']}",
                                            f"Ki ({enzyme['name']}) (μmol/L)",
                                            value=10, step=0.1)
                elif input[f"ki_input_type_{enzyme['id']}"]() == "ic50_convert":
                    return ui.TagList(
                        ui.input_numeric(f"ic50_value_{enzyme['id']}",
                                        f"IC50 ({enzyme['name']}) (μmol/L)",
                                        value=20, step=0.1),
                        ui.input_numeric(f"substrate_conc_{enzyme['id']}",
                                        f"[S] substrate concentration (μmol/L)",
                                        value=1, step=0.1),
                        ui.input_numeric(f"km_value_{enzyme['id']}",
                                        f"Km ({enzyme['name']}) (μmol/L)",
                                        value=10, step=0.1),
                        ui.tags.footer("Ki = IC50 / (1 + [S]/Km)")
                    )
            return None

        @output(id=f"ki_result_{enzyme['id']}")
        @render.text
        def ki_result():
            if input[f"select_{enzyme['id']}"]():
                if input[f"ki_input_type_{enzyme['id']}"]() == "ki_direct":
                    ki_val = input[f"ki_value_{enzyme['id']}"]()
                    return f"Ki: {ki_val} μmol/L"
                elif input[f"ki_input_type_{enzyme['id']}"]() == "ic50_convert":
                    ic50   = input[f"ic50_value_{enzyme['id']}"]()
                    s_conc = input[f"substrate_conc_{enzyme['id']}"]()
                    km     = input[f"km_value_{enzyme['id']}"]()
                    ki_calc = ic50 / (1 + s_conc / km)
                    return f"Calculated Ki: {round(ki_calc, 3)} μmol/L"
            return ""    

    # Helper function to get Ki value (enzyme inhibition)
    def get_ki_value(enzyme_id):
        if input[f"ki_input_type_{enzyme_id}"]() == "ki_direct":
            return input[f"ki_value_{enzyme_id}"]()
        elif input[f"ki_input_type_{enzyme_id}"]() == "ic50_convert":
            ic50 = input[f"ic50_value_{enzyme_id}"]()
            s_conc = input[f"substrate_conc_{enzyme_id}"]()
            km = input[f"km_value_{enzyme_id}"]()
            return ic50 / (1 + s_conc / km)
        return 10  # default value

    # Helper function to get transporter fuinc
    def get_fu_for_transporter(transporter_id):
        try:
            if input[f"use_fuinc_ti_{transporter_id}"]():
                fuinc_val = input[f"fuinc_value_ti_{transporter_id}"]()
                if fuinc_val is not None and fuinc_val > 0:
                    return fuinc_val, "fu,inc (individual)"
        except:
            pass
        return input.fuinc(), "fu,inc (shared)"

    # Register render functions for each enzyme
    for enzyme in all_enzymes_ei_unified:
        make_render_parameters_ei_unified(enzyme)   
        make_ki_input_ui(enzyme)                    

    for enzyme in all_enzymes_tdi:
        make_render_parameters_tdi(enzyme)

    for enzyme in all_enzymes_ed:
        make_render_parameters_ed(enzyme)

    for enzyme in all_enzymes_ne_unified:
        make_render_parameters_ne_unified(enzyme)   

    for transporter in all_transporters_ie:
        make_render_parameters_ti_ie(transporter)

    for transporter in all_transporters_hu:
        make_render_parameters_ti_hu(transporter)

    for transporter in all_transporters_ru:
        make_render_parameters_ti_ru(transporter)

    for transporter in all_transporters_re:
        make_render_parameters_ti_re(transporter)

    # Calculate results for selected enzymes and transporters

    #Competitive inhibition - hepatic
    @output
    @render.ui
    @reactive.event(input.calculate_all_ei)
    def result_table_ei_unified():
        global results_df_ei, results_df_ei_net

        all_enzymes = all_enzymes_ei_reactive()
        selected_enzymes = [enzyme for enzyme in all_enzymes
                            if not enzyme.get("removed", False) and input[f"select_{enzyme['id']}"]()]  
        results_ei     = []
        results_ei_net = []

        for enzyme in selected_enzymes:
            if enzyme.get("is_custom", False):
            #     # Get parameters for custom enzyme using its index
                idx = enzyme.get("custom_index", 0)
                try:
                    ki = input[f"custom_ei_ki_{idx}"]()
                    enzyme_name = input[f"custom_ei_name_{idx}"]() or f"Custom #{idx+1}"
                    probe = input[f"custom_ei_probe_{idx}"]() 
                except Exception as e:
                    print(f"Error getting custom parameters: {e}")
                    continue
            else:
                # Get parameters for predefined enzyme
                ki    = get_ki_value(enzyme['id'])
                enzyme_name = enzyme["name"]
                probe = input[f"probe_{enzyme['id']}"]()
                

            cmax  = get_cmax_umol()
            fup   = input.fup() 
            fu_value, fu_type = get_fu_for_enzyme_ei(enzyme['id'])      

            # ── Hepatic calculation ──────────────────────────────────────────────
            result_hep = cmax * fup / (ki * fu_value)
            alert_hep  = "Risk (R > 0.02)" if result_hep > 0.02 else ""
            note_hep   = f"Ki = {round(ki, 3)} μmol/L, {fu_type} = {round(fu_value, 3)}"

            results_ei.append({
                "Attribute":          f"{enzyme['attribute']} - hepatic",
                "Enzyme/Transporter": enzyme_name,
                "Probe/Clinical Substrate":    probe,  
                "Risk (R)":           round(result_hep, 3),
                "Alert":              alert_hep,
                "Note":               note_hep
            })

            # ── Intestinal calculation (CYP3A4 only) ──────────────────────────────
            if enzyme.get("intestinal", False):
                dose_umol  = 1000000 * input.dose() / input.mw()
                result_int = dose_umol / (250 * ki * fu_value)
                alert_int  = "Risk (R > 10)" if result_int > 10 else ""
                note_int   = (f"Probe substrate: {probe}, Ki = {round(ki, 3)} μmol/L, "
                            f"{fu_type} = {round(fu_value, 3)} (same Ki and {fu_type} as hepatic)")

                results_ei.append({
                    "Attribute":          f"{enzyme['attribute']} - intestinal",
                    "Enzyme/Transporter": enzyme["name"],
                    "Probe/Clinical Substrate":    probe,
                    "Risk (R)":           round(result_int, 3),
                    "Alert":              alert_int,
                    "Note":               note_int
                })

            # Store Ki for net effect calculation (unified ID, no substrate distinction)
            results_ei_net.append({
                "Enzyme/Transporter": enzyme_name,
                "ID":                 enzyme["id"],
                "Ki":                 ki,
                "fu_value":           fu_value,
                "fu_type":            fu_type,
                "intestinal":         enzyme.get("intestinal", False)
            })

        results_df_ei     = pd.DataFrame(results_ei)
        results_df_ei_net = pd.DataFrame(results_ei_net)

        # Apply conditional formatting based on calculation type (hepatic vs intestinal)
        def format_risk(row):
            x         = row["Risk (R)"]
            attribute = row["Attribute"]
            if "intestinal" in attribute:
                # Intestinal threshold: R > 10
                color  = "red"   if x > 10   else "green"
                bold   = "font-weight:bold" if x > 10 else ""
            else:
                # Hepatic threshold: R > 0.02
                color  = "red"   if x > 0.02 else "green"
                bold   = "font-weight:bold" if x > 0.02 else ""
            return f'<span style="color:{color};{bold}">{x}</span>'

        results_df_ei["Risk (R)"] = results_df_ei.apply(format_risk, axis=1)

        return ui.HTML(results_df_ei.to_html(escape=False, index=False))

    #Irreversible CYP inhibition - hepatic    
    @output
    @render.ui
    @reactive.event(input.calculate_all_tdi)
    def result_table_tdi():
        global results_df_tdi, results_df_tdi_net

        # Collect selected enzymes
        all_enzymes = all_enzymes_tdi_reactive()
        selected_enzymes = [enzyme for enzyme in all_enzymes
                            if not enzyme.get("removed", False) and input[f"select_{enzyme['id']}"]()]  
        results_tdi = []
        results_tdi_net = []

        # Perform calculations for each selected enzyme
        for enzyme in selected_enzymes:
            if enzyme.get("is_custom", False):
            #     # Get parameters for custom enzyme using its index
                idx = enzyme.get("custom_index", 0)
                try:
                    ki_tdi = input[f"custom_tdi_ki_tdi_{idx}"]()
                    kdeg_h = input[f"custom_tdi_kdeg_h_{idx}"]()
                    kinact = input[f"custom_tdi_kinact_{idx}"]()
                    enzyme_name = input[f"custom_tdi_name_{idx}"]() or f"Custom #{idx+1}"
                    probe = input[f"custom_tdi_probe_{idx}"]() 
                except Exception as e:
                    print(f"Error getting custom parameters: {e}")  
                    continue
            else:
                # Get parameters for predefined enzyme
                ki_tdi = input[f"ki_tdi_{enzyme['id']}"]()
                kdeg_h = input[f"kdeg_h_{enzyme['id']}"]()
                kinact = input[f"kinact_{enzyme['id']}"]()
                enzyme_name = enzyme["name"]
                probe = input[f"probe_tdi_{enzyme['id']}"]()

            cmax = get_cmax_umol()
            fup = input.fup()
            fu_value, fu_type = get_fu_for_enzyme_tdi(enzyme['id'])
            
            kobs = kinact * 5 * cmax * fup / (ki_tdi * fu_value + 5 * cmax * fup)
            result = (kobs + kdeg_h) / kdeg_h
            if result > 1.25:
                alert = "Risk (R > 1.25)"
            else:
                alert = ""

            note = f"KI = {ki_tdi} μmol/L, Kinact = {kinact}/h, Kdeg,h = {kdeg_h}/h, Kobs = {round(kobs,3)}/h, {fu_type} = {round(fu_value, 3)}"

            results_tdi.append({"Attribute":enzyme["attribute"],"Enzyme/Transporter": enzyme_name, "Probe/Clinical Substrate": probe,"Risk (R)": round(result, 3),"Alert":alert,"Note":note})
            results_tdi_net.append({"Enzyme/Transporter": enzyme_name, "ID": enzyme["id"], "Kdeg,h":kdeg_h,"Kinact":kinact, "KI":ki_tdi,"fu_value":fu_value,"fu_type":fu_type})

        # Convert the results list to a Pandas DataFrame
        results_df_tdi = pd.DataFrame(results_tdi)
        results_df_tdi_net = pd.DataFrame(results_tdi_net)

        # Apply conditional formatting to only the "Risk (R)" column
        results_df_tdi["Risk (R)"] = results_df_tdi["Risk (R)"].apply(
            lambda x: f'<span style="color:{"red" if x > 1.25 else "green"};{"font-weight:bold" if x > 1.25 else ""}">{x}</span>'
        )

        # Convert the DataFrame to HTML
        return ui.HTML(results_df_tdi.to_html(escape=False, index=False))

    #Enzyme induction
    @output
    @render.ui
    @reactive.event(input.calculate_all_ed)
    def result_table_ed():
        global results_df_ed, results_df_ed_net

        # Collect selected enzymes
        all_enzymes = all_enzymes_ed_reactive()
        selected_enzymes = [
            e for e in all_enzymes
            if not e.get("removed", False) and input[f"select_{e['id']}"]()]
        results_ed = []
        results_ed_net = []

        # Perform calculations for each selected enzyme
        for enzyme in selected_enzymes:
            if enzyme.get("is_custom", False):
            #     # Get parameters for custom enzyme using its index
                idx = enzyme.get("custom_index", 0)
                try:
                    emax = input[f"custom_ed_emax_{idx}"]()
                    ec50 = input[f"custom_ed_ec50_{idx}"]()
                    enzyme_name = input[f"custom_ed_name_{idx}"]() or f"Custom #{idx+1}"
                except Exception as e:
                    print(f"Error getting custom parameters: {e}")
                    continue
            else:
                # Get parameters for predefined enzyme
                emax = input[f"emax_{enzyme['id']}"]()
                ec50 = input[f"ec50_{enzyme['id']}"]()
                enzyme_name = enzyme["name"]

            cmax = get_cmax_umol()
            fup = input.fup()
            fu_value, fu_type = get_fu_for_enzyme_ed(enzyme['id'])
            d = input.d_factor()

            result = 1 / (1 + (10 * d * emax * cmax * fup) / (ec50 * fu_value + 10 * cmax * fup))

            if result < 0.2:
                alert = "Strong inducer (R < 0.2)"
            elif 0.2 <= result < 0.5:
                alert = "Moderate inducer (0.2 ≤ R < 0.5)"
            elif 0.5 <= result < 0.8:
                alert = "Weak inducer (0.5 ≤ R < 0.8)"
            else:
                alert = ""

            note = f"Emax = {emax}, EC50 = {ec50} μmol/L, {fu_type} = {round(fu_value, 3)}"

            results_ed.append({"Attribute":enzyme["attribute"],"Enzyme/Transporter": enzyme_name, "Probe/Clinical Substrate": "-","Risk (R)": round(result, 3),"Alert":alert,"Note":note})
            results_ed_net.append({"Enzyme/Transporter": enzyme_name, "ID": enzyme["id"],"Emax":emax,"EC50":ec50,"fu_value":fu_value,"fu_type":fu_type})

        # Convert the results list to a Pandas DataFrame
        results_df_ed = pd.DataFrame(results_ed)
        results_df_ed_net = pd.DataFrame(results_ed_net)

        # Apply conditional formatting to only the "Risk (R)" column
        results_df_ed["Risk (R)"] = results_df_ed["Risk (R)"].apply(
            lambda x: f'<span style="color:{"red" if x < 0.2 else "orange" if 0.2 <= x < 0.5 else "blue" if 0.5 <= x < 0.8 else "green"};{"font-weight:bold" if x < 0.8 else ""}">{x}</span>'
           )

        # Convert the DataFrame to HTML
        return ui.HTML(results_df_ed.to_html(escape=False, index=False))


    #Net effect
    def safe_lookup(df, filter_col, filter_val, target_col):
        try:
            if df is None or len(df) == 0:
                return None
            return df.loc[df[filter_col] == filter_val, target_col].values[0]
        except (NameError, KeyError, IndexError, AttributeError):
            return None

        # return ag, bg, cg, ah, bh, ch
    def calc_abcgh(ki, kinact, kdeg_h, kdeg_g, ki_tdi, emax, ec50,
               i_g, i_h, fu_ei, fu_tdi, fu_ed, d):
        ag = 1 if ki is None else 1 / (1 + (i_g / (fu_ei * ki)))
        bg = 1 if any(v is None for v in [kinact, kdeg_g, ki_tdi]) \
            else kdeg_g / (kdeg_g + (i_g * kinact / (i_g + fu_tdi * ki_tdi)))
        cg = 1 if any(v is None for v in [emax, ec50]) \
            else 1 + (d * emax * i_g) / (i_g + fu_ed * ec50)

        ah = 1 if ki is None else 1 / (1 + (i_h / (fu_ei * ki)))
        bh = 1 if any(v is None for v in [kinact, kdeg_h, ki_tdi]) \
            else kdeg_h / (kdeg_h + (i_h * kinact / (i_h + fu_tdi * ki_tdi)))
        ch = 1 if any(v is None for v in [emax, ec50]) \
            else 1 + (d * emax * i_h) / (i_h + fu_ed * ec50)

        return ag, bg, cg, ah, bh, ch

    def calc_aucr(ag, bg, cg, ah, bh, ch, fg_victim, fm):
        result_ei    = (1 / (ag * bg  * 1  * (1 - fg_victim) + fg_victim)) * (1 / (ah * bh * 1  * fm + (1 - fm)))
        result_ed    = (1 / (1  * 1 * cg * (1 - fg_victim) + fg_victim)) * (1 / (1  * 1  * ch * fm + (1 - fm)))
        result_total = (1 / (ag * bg * cg * (1 - fg_victim) + fg_victim)) * (1 / (ah * bh * ch * fm + (1 - fm)))
        return result_ei, result_ed, result_total

    def get_alerts(result_ei, result_ed, result_total):
        alert_ei    = "Risk (R > 1.25)" if result_ei > 1.25 else ""
        alert_ed    = "Risk (R < 0.8)"  if result_ed < 0.8  else ""
        alert_total = "Risk (R < 0.8)"  if result_total < 0.8 else ("Risk (R > 1.25)" if result_total > 1.25 else "")
        return alert_ei, alert_ed, alert_total

    def get_notes(i_g, i_h, ag, bg, cg, ah, bh, ch, fg_victim, fm, ki, kinact, ki_tdi, fu_ei_val, fu_tdi_val, kdeg_g, kdeg_h, d, emax, ec50, fu_ed_val):
        note_ei    = f"Ag={round(ag,3)}, Bg={round(bg,3)}, Cg=1, Fg,victim={fg_victim}, Ah={round(ah,3)}, Bh={round(bh,3)}, Ch=1, fm={fm}: [I]g={round(i_g,3)}, [I]h={round(i_h,3)}, Ki={ki}, fu,mic(A)={fu_ei_val}, Kinact={kinact}, KI={ki_tdi}, fu_mic(B)={fu_tdi_val}, Kdeg,g={kdeg_g}, Kdeg,h={kdeg_h}"
        note_ed    = f"Ag=1, Bg=1, Cg={round(cg,3)}, Fg,victim={fg_victim}, Ah=1, Bh=1, Ch={round(ch,3)}, fm={fm}: [I]g={round(i_g,3)}, [I]h={round(i_h,3)}, d={d}, Emax={emax}, EC50={ec50}, fu,hep(C)={fu_ed_val}"
        note_total = f"Ag={round(ag,3)}, Bg={round(bg,3)}, Cg={round(cg,3)}, Fg,victim={fg_victim}, Ah={round(ah,3)}, Bh={round(bh,3)}, Ch={round(ch,3)}, fm={fm}: [I]g={round(i_g,3)}, [I]h={round(i_h,3)}, Ki={ki}, fu,mic(A)={fu_ei_val}, Kinact={kinact}, KI={ki_tdi}, fu_mic(B)={fu_tdi_val}, Kdeg,g={kdeg_g}, Kdeg,h={kdeg_h}, d={d}, Emax={emax}, EC50={ec50}, fu,hep(C)={fu_ed_val}"
        return note_ei, note_ed, note_total

    def append_ne_results(results_ne_ei, results_ne_ed, results_ne_total,
                          enzyme_name, substrate_name, fm, fg_victim,
                          ag, bg, cg, ah, bh, ch, i_g, i_h, 
                          ki, kinact, ki_tdi, fu_ei_val, fu_tdi_val, kdeg_g, kdeg_h, d, emax, ec50, fu_ed_val):
        result_ei, result_ed, result_total = calc_aucr(ag, bg, cg, ah, bh, ch, fg_victim, fm)
        alert_ei, alert_ed, alert_total    = get_alerts(result_ei, result_ed, result_total)
        note_ei, note_ed, note_total       = get_notes(i_g, i_h, ag, bg, cg, ah, bh, ch, fg_victim, fm, ki, kinact, ki_tdi, fu_ei_val, fu_tdi_val, kdeg_g, kdeg_h, d, emax, ec50, fu_ed_val)

        results_ne_ei.append({
            "Attribute": f'{enzyme["attribute"]} (inhibition only)',
            "Enzyme/Transporter": enzyme_name,
            "Probe/Clinical Substrate": substrate_name,
            "Risk (R)": round(result_ei, 3),
            "Alert": alert_ei,
            "Note": note_ei
        })
        results_ne_ed.append({
            "Attribute": f'{enzyme["attribute"]} (induction only)',
            "Enzyme/Transporter": enzyme_name,
            "Probe/Clinical Substrate": substrate_name,
            "Risk (R)": round(result_ed, 3),
            "Alert": alert_ed,
            "Note": note_ed
        })
        results_ne_total.append({
            "Attribute": f'{enzyme["attribute"]} (AUCR)',
            "Enzyme/Transporter": enzyme_name,
            "Probe/Clinical Substrate": substrate_name,
            "Risk (R)": round(result_total, 3),
            "Alert": alert_total,
            "Note": note_total
        })

    @output
    @render.ui
    @reactive.event(input.calculate_all_ne)
    def result_table_ne_unified():
        global results_df_ne, results_df_ei_net, results_df_tdi_net, results_df_ed_net

        all_enzymes = all_enzymes_ne_reactive()
        selected_enzymes = [
            enzyme for enzyme in all_enzymes
            if not enzyme.get("removed", False) and input[f"select_{enzyme['id']}"]()
        ]

        if not selected_enzymes:
            return ui.HTML("<p style='color:orange;'>Please select at least one enzyme for Net Effect calculation.</p>")

        results_ne_ei    = []  # inhibition-only results
        results_ne_ed    = []  # induction-only results
        results_ne_total = []  # total net effect results

        # --- Retrieve common pharmacokinetic input parameters ---
        fup   = input.fup()
        cmax  = get_cmax_umol()
        dose  = input.dose() / input.mw()   # dose in mmol
        fa    = input.fa()
        fg    = input.fg()
        ka    = input.ka()
        qen   = input.qen()
        qh    = input.qh()
        rb    = input.rb()
        d = input.d_factor()
        
        try:
            fumic = calculate_fumic()
        except Exception as e:
            fumic = 1.0
        
        try:
            fuhep = calculate_fuhep()
        except Exception as e:
            fuhep = 1.0


        # --- Calculate intestinal and hepatic inhibitor concentrations ---
        i_g = fa * ka * dose * 1000 / qen
        i_h = fup * (cmax + (fa * fg * ka * dose * 1000) / qh / rb)

        # Main loop - FULLY UNIFIED for ALL enzymes including CYP3A4
        for enzyme in selected_enzymes:

            # Default values
            enzyme_name = enzyme.get("name", "Unknown")
            substrate_name = "Unknown"
            fm = 0.5
            fg_victim = 1.0
            kdeg_g = None

            if enzyme.get("is_custom", False):
                idx = enzyme.get("custom_index", 0)
                try:
                    enzyme_name = input[f"custom_ne_name_{idx}"]() or f"Custom #{idx+1}"
                    substrate_name = input[f"custom_ne_substrate_{idx}"]() or "Custom substrate"
                    fm = input[f"custom_ne_fm_{idx}"]()
                    fg_victim = input[f"custom_ne_fg_{idx}"]()
                except Exception as e:
                    print(f"Error getting custom NE parameters: {e}")
                    continue
            else:
                try:
                    enzyme_name = enzyme["name"]
                    substrate_name = input[f"substrate_name_{enzyme['id']}"]() or enzyme.get("substrate", "Unknown")
                    fm = input[f"fm_{enzyme['id']}"]()
                    fg_victim = input[f"fg_victim_{enzyme['id']}"]()
                except Exception as e:
                    continue

            # Retrieve induction parameters from enzyme induction results
            emax = safe_lookup(results_df_ed_net, "Enzyme/Transporter", enzyme_name, "Emax")
            ec50 = safe_lookup(results_df_ed_net, "Enzyme/Transporter", enzyme_name, "EC50")

            # --- Look up inhibition parameters by enzyme name ---
            ki     = safe_lookup(results_df_ei_net,  "Enzyme/Transporter", enzyme_name, "Ki")
            kinact = safe_lookup(results_df_tdi_net, "Enzyme/Transporter", enzyme_name, "Kinact")
            ki_tdi = safe_lookup(results_df_tdi_net, "Enzyme/Transporter", enzyme_name, "KI")
            
            kdeg_h   = safe_lookup(results_df_tdi_net, "Enzyme/Transporter", enzyme_name, "Kdeg,h")
            if kdeg_h is None:
                kdeg_h = next(
                    (e["kdeg_h"] for e in all_enzymes_tdi if e["name"] == enzyme_name),
                    None  
                )
            
            kdeg_g = get_kdeg_g_for_enzyme_ne(enzyme["id"], enzyme_name, kdeg_h)

            # --- fu lookup per mechanism ---
            fu_ei_val = safe_lookup(results_df_ei_net, "Enzyme/Transporter", enzyme_name, "fu_value")
            if not isinstance(fu_ei_val, (int, float)):
                fu_ei_val = fumic

            fu_tdi_val = safe_lookup(results_df_tdi_net, "Enzyme/Transporter", enzyme_name, "fu_value")
            if not isinstance(fu_tdi_val, (int, float)):
                fu_tdi_val = fumic
            
            fu_ed_val = safe_lookup(results_df_ed_net, "Enzyme/Transporter", enzyme_name, "fu_value")
            if not isinstance(fu_ed_val, (int, float)):
                fu_ed_val = fuhep

            # Calculate the six mechanistic factors (same function for all enzymes)
            ag, bg, cg, ah, bh, ch = calc_abcgh(
                ki, kinact, kdeg_h, kdeg_g, ki_tdi,
                emax, ec50, i_g, i_h, fu_ei_val, fu_tdi_val, fu_ed_val, d
            )

            # Append results for the primary substrate
            append_ne_results(
                results_ne_ei, results_ne_ed, results_ne_total,
                enzyme_name, substrate_name, fm, fg_victim,
                ag, bg, cg, ah, bh, ch, i_g, i_h, 
                ki, kinact, ki_tdi, fu_ei_val, fu_tdi_val, kdeg_g, kdeg_h, d, emax, ec50, fu_ed_val
            )

        # --- Combine all result lists and sort ---
        combined_results = results_ne_ei + results_ne_ed + results_ne_total

        if not combined_results:
            return ui.HTML("<p style='color:orange;'>No results calculated. Please select enzymes and ensure input parameters are valid.</p>")
        results_df_ne_beforesort = pd.DataFrame(combined_results)
        results_df_ne = results_df_ne_beforesort.sort_values(by=["Enzyme/Transporter", "Probe/Clinical Substrate", "Attribute"])

        # --- Apply conditional color formatting to Risk (R) column ---
        results_df_ne["Risk (R)"] = results_df_ne["Risk (R)"].apply(
            lambda x: f'<span style="color:{"red" if x < 0.8 or x > 1.25 else "green"};'
                      f'{"font-weight:bold" if x < 0.8 or x > 1.25 else ""}">{x}</span>'
        )

        return ui.HTML(results_df_ne.to_html(escape=False, index=False))

    #Transporter inhibition - intestinal efflux    
    @output
    @render.ui
    @reactive.event(input.calculate_all_ti_ie)
    def result_table_ti_ie():
        global results_df_ti_ie

        # Collect selected transporters
        all_transporters = all_transporters_ie_reactive()
        selected_transporters = [transporter for transporter in all_transporters
                                if not transporter.get("removed", False) and input[f"select_{transporter['id']}"]()]  
        results_ti_ie = []

        # Perform calculations for each selected transporter
        for transporter in selected_transporters:
            if transporter.get("is_custom", False):
            #     # Get parameters for custom transporter using its index
                idx = transporter.get("custom_index", 0)
                try:
                    ic50 = input[f"custom_ti_ie_ic50_{idx}"]()
                    transporter_name = input[f"custom_ti_ie_name_{idx}"]() or f"Custom #{idx+1}"
                except Exception as e:
                    print(f"Error getting custom parameters: {e}")
                    continue
            else:
                # Get parameters for predefined transporter
                ic50 = input[f"ic50_{transporter['id']}"]()
                transporter_name = transporter["name"]

            dose = 1000000 * input.dose() / input.mw()
            fu_value, fu_type = get_fu_for_transporter(transporter['id'])
            result = dose / (250 * ic50 * fu_value)
            if result > 10:
                alert = "Risk (R > 10)"
            else:
                alert = ""

            note = f"IC50 or Ki = {ic50} μmol/L, {fu_type} = {round(fu_value, 3)}"

            results_ti_ie.append({"Attribute":transporter["attribute"],"Enzyme/Transporter": transporter_name, "Probe/Clinical Substrate": "-", "Risk (R)": round(result, 3),"Alert":alert,"Note":note})

        # Convert the results list to a Pandas DataFrame
        results_df_ti_ie = pd.DataFrame(results_ti_ie)

        # Apply conditional formatting to only the "Risk (R)" column
        results_df_ti_ie["Risk (R)"] = results_df_ti_ie["Risk (R)"].apply(
            lambda x: f'<span style="color:{"red" if x > 10 else "green"};{"font-weight:bold" if x > 10 else ""}">{x}</span>'
        )

        # Convert the DataFrame to HTML
        return ui.HTML(results_df_ti_ie.to_html(escape=False, index=False))


    #Transporter inhibition - hepatic uptake    
    @output
    @render.ui
    @reactive.event(input.calculate_all_ti_hu)
    def result_table_ti_hu():
        global results_df_ti_hu

        # Collect selected transporters
        all_transporters = all_transporters_hu_reactive()
        selected_transporters = [transporter for transporter in all_transporters
                                if not transporter.get("removed", False) and input[f"select_{transporter['id']}"]()]
        results_ti_hu = []

        # Perform calculations for each selected transporter
        for transporter in selected_transporters:
            if transporter.get("is_custom", False):
            #     # Get parameters for custom transporter using its index
                idx = transporter.get("custom_index", 0)
                try:
                    ic50 = input[f"custom_ti_hu_ic50_{idx}"]()
                    transporter_name = input[f"custom_ti_hu_name_{idx}"]() or f"Custom #{idx+1}"
                except Exception as e:
                    print(f"Error getting custom parameters: {e}")
                    continue
            else:
                # Get parameters for predefined transporter
                ic50 = input[f"ic50_{transporter['id']}"]()
                transporter_name = transporter["name"]

            fup = input.fup()
            fu_value, fu_type = get_fu_for_transporter(transporter['id'])
            cmax = get_cmax_umol()
            dose = input.dose() / input.mw()
            fa = input.fa()
            fg = input.fg()
            ka = input.ka()
            qh = input.qh()
            rb = input.rb()
            result = (cmax + (fa * fg * ka * dose * 1000 / qh / rb)) * fup / (ic50 * fu_value)
            if result > 0.1:
                alert = "Risk (R > 0.1)"
            else:
                alert = ""

            note = f"IC50 or Ki = {ic50} μmol/L, {fu_type} = {round(fu_value, 3)}"

            results_ti_hu.append({"Attribute":transporter["attribute"],"Enzyme/Transporter": transporter_name, "Probe/Clinical Substrate": "-", "Risk (R)": round(result, 3),"Alert":alert,"Note":note})

        # Convert the results list to a Pandas DataFrame
        results_df_ti_hu = pd.DataFrame(results_ti_hu)

        # Apply conditional formatting to only the "Risk (R)" column
        results_df_ti_hu["Risk (R)"] = results_df_ti_hu["Risk (R)"].apply(
            lambda x: f'<span style="color:{"red" if x > 0.1 else "green"};{"font-weight:bold" if x > 0.1 else ""}">{x}</span>'
        )

        # Convert the DataFrame to HTML
        return ui.HTML(results_df_ti_hu.to_html(escape=False, index=False))


    #Transporter inhibition - renal uptake    
    @output
    @render.ui
    @reactive.event(input.calculate_all_ti_ru)
    def result_table_ti_ru():
        global results_df_ti_ru

        # Collect selected transporters
        all_transporters = all_transporters_ru_reactive()
        selected_transporters = [transporter for transporter in all_transporters
                                if not transporter.get("removed", False) and input[f"select_{transporter['id']}"]()]  
        results_ti_ru = []

        # Perform calculations for each selected transporter
        for transporter in selected_transporters:
            if transporter.get("is_custom", False):
            #     # Get parameters for custom transporter using its index
                idx = transporter.get("custom_index", 0)
                try:
                    ic50 = input[f"custom_ti_ru_ic50_{idx}"]()
                    transporter_name = input[f"custom_ti_ru_name_{idx}"]() or f"Custom #{idx+1}"
                except Exception as e:
                    print(f"Error getting custom parameters: {e}")
                    continue
            else:
                # Get parameters for predefined transporter
                ic50 = input[f"ic50_{transporter['id']}"]()
                transporter_name = transporter["name"]
            
            fup = input.fup()
            fu_value, fu_type = get_fu_for_transporter(transporter['id'])
            cmax = get_cmax_umol()
            result = cmax * fup / (ic50 * fu_value)
            if result > 0.1:
                alert = "Risk (R > 0.1)"
            else:
                alert = ""

            note = f"IC50 or Ki = {ic50} μmol/L, {fu_type} = {round(fu_value, 3)}"

            results_ti_ru.append({"Attribute":transporter["attribute"],"Enzyme/Transporter": transporter_name, "Probe/Clinical Substrate": "-", "Risk (R)": round(result, 3),"Alert":alert,"Note":note})

        # Convert the results list to a Pandas DataFrame
        results_df_ti_ru = pd.DataFrame(results_ti_ru)

        # Apply conditional formatting to only the "Risk (R)" column
        results_df_ti_ru["Risk (R)"] = results_df_ti_ru["Risk (R)"].apply(
            lambda x: f'<span style="color:{"red" if x > 0.1 else "green"};{"font-weight:bold" if x > 0.1 else ""}">{x}</span>'
        )

        # Convert the DataFrame to HTML
        return ui.HTML(results_df_ti_ru.to_html(escape=False, index=False))


    #Transporter inhibition - renal efflux    
    @output
    @render.ui
    @reactive.event(input.calculate_all_ti_re)
    def result_table_ti_re():
        global results_df_ti_re

        # Collect selected enzymes
        all_transporters = all_transporters_re_reactive()
        selected_transporters = [transporter for transporter in all_transporters
                                if not transporter.get("removed", False) and input[f"select_{transporter['id']}"]()]  
        results_ti_re = []

        # Perform calculations for each selected transporter
        for transporter in selected_transporters:
            if transporter.get("is_custom", False):
            #     # Get parameters for custom transporter using its index
                idx = transporter.get("custom_index", 0)
                try:
                    ic50 = input[f"custom_ti_re_ic50_{idx}"]()
                    transporter_name = input[f"custom_ti_re_name_{idx}"]() or f"Custom #{idx+1}"
                except Exception as e:
                    print(f"Error getting custom parameters: {e}")
                    continue
            else:
                # Get parameters for predefined transporter
                ic50 = input[f"ic50_{transporter['id']}"]()
                transporter_name = transporter["name"]

            fup = input.fup()
            fu_value, fu_type = get_fu_for_transporter(transporter['id'])
            cmax = get_cmax_umol()
            result = cmax * fup / (ic50 * fu_value)
            if result > 0.02:
                alert = "Risk (R > 0.02)"
            else:
                alert = ""

            note = f"IC50 or Ki = {ic50} μmol/L, {fu_type} = {round(fu_value, 3)}"

            results_ti_re.append({"Attribute":transporter["attribute"],"Enzyme/Transporter": transporter_name, "Probe/Clinical Substrate": "-", "Risk (R)": round(result, 3),"Alert":alert,"Note":note})

        # Convert the results list to a Pandas DataFrame
        results_df_ti_re = pd.DataFrame(results_ti_re)

        # Apply conditional formatting to only the "Risk (R)" column
        results_df_ti_re["Risk (R)"] = results_df_ti_re["Risk (R)"].apply(
            lambda x: f'<span style="color:{"red" if x > 0.02 else "green"};{"font-weight:bold" if x > 0.02 else ""}">{x}</span>'
        )

        # Convert the DataFrame to HTML
        return ui.HTML(results_df_ti_re.to_html(escape=False, index=False))


    # Summary
    @output
    @render.ui
    @reactive.event(input.summary)
    def summary_table(): 
        global unified_df1, unified_df2, unified_df3, unified_df4, user_memo_df

        # List of DataFrame names for each table
        df_names1 = ["results_df_ei", "results_df_tdi"]
        df_names2 = ["results_df_ed"]
        df_names3 = ["results_df_ne"]
        df_names4 = ["results_df_ti_ie", "results_df_ti_hu", "results_df_ti_ru", "results_df_ti_re"]
        
        # Function to concatenate DataFrames from a list of names
        def concatenate_dfs(df_names):
            dfs = []
            for name in df_names:
                try:
                    df = eval(name)
                    if df is not None and not df.empty:
                        dfs.append(df)
                except NameError:
                    pass
            return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

        # Create three separate DataFrames
        unified_df1 = concatenate_dfs(df_names1)
        unified_df2 = concatenate_dfs(df_names2)
        unified_df3 = concatenate_dfs(df_names3)
        unified_df4 = concatenate_dfs(df_names4)

        # Convert the DataFrames to HTML
        html_table1 = unified_df1.to_html(escape=False, index=False)
        html_table2 = unified_df2.to_html(escape=False, index=False)
        html_table3 = unified_df3.to_html(escape=False, index=False)
        html_table4 = unified_df4.to_html(escape=False, index=False)

        # Combine the HTML tables with separators
        combined_html = (
            f"<h2 style='color:#4B0082'>Basic Static Models</h2>"
            f"<h3>Enzyme Inhibition</h3>{html_table1}<br>"
            f"<h3>Enzyme Induction</h3>{html_table2}<br>"
            f"<h3>Transporter Inhibition</h3>{html_table4}<br><br>"
            f"<h2 style='color:#006400'>Mechanistic Static Model</h2>"
            f"<h3>Net Effect (AUCR)</h3>{html_table3}<br>"
        )

        # Convert user_memo to DataFrame
        user_memo = input.user_memo()
        user_memo_lines = user_memo.split('\n')  # Split the text into lines
        user_memo_data = [line.split(',') for line in user_memo_lines]  # Split each line into columns
        user_memo_df = pd.DataFrame(user_memo_data[1:], columns=user_memo_data[0])  # Create DataFrame

        return ui.HTML(combined_html)

    # Output excel file
    # Function to strip HTML tags and extract styles
    def extract_text_and_style(html):
        soup = BeautifulSoup(html, "html.parser")
        text = soup.get_text()
        style = {}
        if soup.span:
            style['bold'] = 'font-weight:bold' in soup.span.get('style', '')
            style['color'] = soup.span.get('style', '').split('color:')[-1].split(';')[0] if 'color:' in soup.span.get('style', '') else None
            style['bgcolor'] = soup.span.get('style', '').split('background-color:')[-1].split(';')[0] if 'background-color:' in soup.span.get('style', '') else None
        return text, style

    # Function to convert color to aRGB hex format
    def to_argb_hex(color):
        if color.startswith('#'):
            color = color[1:]
        if len(color) == 6:
            color = 'FF' + color  # Add alpha value
        return color.upper()

    # Mapping of named colors to hex values
    named_colors = {
        'red': 'FF0000',
        'green': '00FF00',
        'blue': '0000FF',
        'black': '000000',
        'white': 'FFFFFF',
        'yellow': 'FFFF00',
        'cyan': '00FFFF',
        'magenta': 'FF00FF',
        'orange': 'FFA500'
    }

    # Function to validate and convert color
    def validate_and_convert_color(color):
        if color:
            if color in named_colors:
                color = named_colors[color]
            elif color.startswith('rgb'):
                # Convert rgb(r, g, b) to hex
                color = color.replace('rgb(', '').replace(')', '')
                r, g, b = map(int, color.split(','))
                color = f'FF{r:02X}{g:02X}{b:02X}'
            elif color.startswith('#'):
                color = to_argb_hex(color)
            return color
        return None

    @output
    @render.download(filename=f"statDDI_{datetime.now().strftime('%Y%m%d')}.xlsx")
    def handle_download_xlsx():
        def excel_content():
            excel_buffer = io.BytesIO()
            workbook = Workbook()
            worksheet = workbook.active

            drug_product = input.cmp()
            mw = input.mw()
            cmax = input.cmax()
            cmax_unit = input.cmax_unit()
            fup = input.fup()
            dose = input.dose()
            ka = input.ka()
            fa = input.fa()
            fg = input.fg()
            qh = input.qh()
            rb = input.rb()
            d_factor = input.d_factor()
            qen = input.qen()
            fuinc = input.fuinc()

            cmax_display = "ng/mL" if cmax_unit == "ng_ml" else "μmol/L"
            profile_text = f'Compound: {drug_product}'
            input_values = (
                f'MW = {mw} g/mol, Cmax = {cmax} {cmax_display}, fu,p = {fup}, '
                f'Dose = {dose} mg, ka = {ka}/h, Fa = {fa}, Fg = {fg}, '
                f'Qh = {qh} L/h, Qen = {qen} L/h, Rb = {rb}, '
                f'd factor = {d_factor}, fu,inc = {fuinc}'
            )

            cell = worksheet.cell(row=1, column=1, value=profile_text)
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(wrap_text=True)

            cell = worksheet.cell(row=2, column=1, value=input_values)
            cell.alignment = Alignment(wrap_text=False)

            # ── helper: write a section heading ──────────────────────────────────
            def write_section_heading(title, start_row, color_hex="4B0082"):
                cell = worksheet.cell(row=start_row, column=1, value=title)
                cell.font = Font(bold=True, size=14, color=f"FF{color_hex}")
                return start_row + 1

            # ── helper: write a DataFrame with a sub-title ────────────────────────
            def write_df_with_title(df, title, start_row):
                worksheet.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=start_row + 1, column=col_num, value=column_title)
                    cell.font = Font(bold=True)
                for r_idx, row in enumerate(df.itertuples(index=False), start_row + 2):
                    for c_idx, value in enumerate(row, 1):
                        if isinstance(value, str) and '<span' in value:
                            text, style = extract_text_and_style(value)
                            cell = worksheet.cell(row=r_idx, column=c_idx, value=text)
                            font_args = {}
                            if style.get('bold'):
                                font_args['bold'] = True
                            if style.get('color'):
                                color = validate_and_convert_color(style['color'])
                                if color:
                                    font_args['color'] = Color(rgb=color)
                            cell.font = Font(**font_args)
                        else:
                            worksheet.cell(row=r_idx, column=c_idx, value=value)
                return start_row + len(df) + 3

            next_row = 4

            # ── BASIC STATIC MODELS ───────────────────────────────────────────────
            next_row = write_section_heading("Basic Static Models", next_row, color_hex="4B0082")
            next_row = write_df_with_title(unified_df1, "Enzyme Inhibition", next_row)
            next_row = write_df_with_title(unified_df2, "Enzyme Induction", next_row)
            next_row = write_df_with_title(unified_df4, "Transporter Inhibition", next_row)

            # ── MECHANISTIC STATIC MODEL ──────────────────────────────────────────
            next_row = write_section_heading("Mechanistic Static Model", next_row, color_hex="006400")
            next_row = write_df_with_title(unified_df3, "Net Effect (AUCR)", next_row)

            # ── USER MEMO ─────────────────────────────────────────────────────────
            write_df_with_title(user_memo_df, "User Memo", next_row)

            # Adjust column widths
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column].width = max_length + 2

            worksheet.column_dimensions['A'].width = 50

            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            return excel_buffer.getvalue()

        try:
            yield excel_content()
        except asyncio.CancelledError:
            print("Download was cancelled.")

    @output
    @render.download(filename=f"statDDI_{datetime.now().strftime('%Y%m%d')}.pdf")
    def handle_download_pdf():
        def pdf_content():
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, pagesize=A4,
                leftMargin=10*mm, rightMargin=10*mm,
                topMargin=10*mm, bottomMargin=10*mm
            )
            elements = []

            styles = getSampleStyleSheet()
            title_style = styles['Title']
            normal_style = styles['Normal']
            bold_style = ParagraphStyle(
                name='Bold', parent=styles['Normal'], fontName='Helvetica-Bold'
            )

            # ── Section heading styles ────────────────────────────────────────────
            section_basic = ParagraphStyle(
                name='SectionBasic', parent=styles['Heading1'],
                textColor=colors.HexColor('#4B0082'),   # dark purple
                fontSize=13, spaceAfter=4
            )
            section_mech = ParagraphStyle(
                name='SectionMech', parent=styles['Heading1'],
                textColor=colors.HexColor('#006400'),   # dark green
                fontSize=13, spaceAfter=4
            )

            # Header
            elements.append(Paragraph('Summary report (Static DDI Risk Calculator)', title_style))
            elements.append(Spacer(1, 0.5*cm))

            # Input parameters
            drug_product = input.cmp()
            mw = input.mw()
            cmax = input.cmax()
            cmax_unit = input.cmax_unit()
            fup = input.fup()
            dose = input.dose()
            ka = input.ka()
            fa = input.fa()
            fg = input.fg()
            qh = input.qh()
            rb = input.rb()
            d_factor = input.d_factor()
            qen = input.qen()
            fuinc = input.fuinc()

            cmax_display = "ng/mL" if cmax_unit == "ng_ml" else "μmol/L"
            elements.append(Paragraph(f'Compound: {drug_product}', bold_style))
            elements.append(Paragraph(
                f'MW = {mw} g/mol, Cmax = {cmax} {cmax_display}, fu,p = {fup}, '
                f'Dose = {dose} mg, ka = {ka}/h, Fa = {fa}, Fg = {fg}, '
                f'Qh = {qh} L/h, Qen = {qen} L/h, Rb = {rb}, '
                f'd factor = {d_factor}, fu,inc = {fuinc}',
                normal_style
            ))
            elements.append(Spacer(1, 0.5*cm))

            # ── helper: write one DataFrame as a PDF table ────────────────────────
            def write_df_to_pdf(df, title,
                                add_title=True, add_background=True, add_grid=True):
                if add_title:
                    elements.append(Paragraph(title, bold_style))
                if df.empty:
                    elements.append(Paragraph("No data available", normal_style))
                    elements.append(Spacer(1, 1*cm))
                    return
                data = [df.columns.tolist()] + df.values.tolist()
                formatted_data = []
                for row in data:
                    formatted_row = []
                    for cell in row:
                        cell_html = str(cell)
                        soup = BeautifulSoup(cell_html, 'html.parser')
                        text_only = soup.get_text()
                        formatted_row.append(Paragraph(text_only, normal_style))
                    formatted_data.append(formatted_row)
                table = Table(formatted_data, repeatRows=1)
                table_style_cmds = [
                    ('ALIGN',        (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN',       (0, 0), (-1, -1), 'TOP'),
                    ('FONTNAME',     (0, 0), (-1,  0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING',(0, 0), (-1,  0), 12),
                ]
                if add_background:
                    table_style_cmds += [
                        ('BACKGROUND', (0, 0), (-1,  0), colors.lightgrey),
                        ('TEXTCOLOR',  (0, 0), (-1,  0), colors.black),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ]
                if add_grid:
                    table_style_cmds.append(('GRID', (0, 0), (-1, -1), 1, colors.black))
                table.setStyle(TableStyle(table_style_cmds))
                elements.append(table)
                elements.append(Spacer(1, 0.5*cm))

            # ── BASIC STATIC MODELS ───────────────────────────────────────────────
            elements.append(Paragraph("Basic Static Models", section_basic))
            elements.append(Spacer(1, 0.2*cm))
            write_df_to_pdf(unified_df1, "Enzyme Inhibition")
            write_df_to_pdf(unified_df2, "Enzyme Induction")
            write_df_to_pdf(unified_df4, "Transporter Inhibition")

            # ── MECHANISTIC STATIC MODEL ──────────────────────────────────────────
            elements.append(Paragraph("Mechanistic Static Model", section_mech))
            elements.append(Spacer(1, 0.2*cm))
            write_df_to_pdf(unified_df3, "Net Effect (AUCR)")

            # ── USER MEMO ─────────────────────────────────────────────────────────
            write_df_to_pdf(
                user_memo_df, "User Memo",
                add_title=True, add_background=False, add_grid=False
            )

            doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
            buffer.seek(0)
            return buffer.getvalue()

        page_count = [0]

        def add_page_number(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            page_count[0] += 1
            page_width = doc.pagesize[0]
            canvas.drawCentredString(page_width / 2.0, 10*mm, f"Page {page_count[0]}")
            canvas.restoreState()

        try:
            yield pdf_content()
        except asyncio.CancelledError:
            print("Download was cancelled.")

    # Glossary
    @output
    @render.table
    def glossary_table():
        return glossary_df

    results = reactive.Value({
        "enzyme_inhibition": None,
        "enzyme_induction": None,
        "net_effect": None,
        "transporter_inhibition": None,

    })


# Create the app
app = App(app_ui, server)
